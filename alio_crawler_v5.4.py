# -*- coding: utf-8 -*-
# ──────────────────────────────────────────────────
# 프로그램명: ALIO Crawler (알리오 크롤러)
# 버전: 5.4.1
# 저작자: 허재영
# 창작연도: 2025
# 최종 수정일: 2026-04-26
# Copyright (c) 2025-2026 허재영. All rights reserved.
# ──────────────────────────────────────────────────
"""
ALIO 공공기관 경영정보 공개시스템 - 항목별 공시 크롤링 시스템 v5
- 기관유형, 주무부처, 지역별 필터링 기능
- 내부규정 크롤링 기능 추가
- 버전: 5.4.1 (2026년 4월 기준)

변경사항 (v5.4 → v5.4.1):
- 게시판형(reportYn=N) 항목 첨부파일 다운로드 기능 추가
- 감사원 지적사항/국회지적사항/주무부처 지적사항/입찰공고 등 12개 항목 실제 PDF 수집 가능
- itemBoard{reportFormNo}.do HTML 파싱 + /upload{spath}{sfile} 직접 다운로드 패턴
- v5.3 환경법규 위반현황(envlaw)에도 첨부파일 다운로드 자동 추가 (부수 개선)

변경사항 (v5.3 → v5.4):
- 알리오 항목별공시 전 항목(약 100여개) 자동 수집 기능 추가 (rootNo 범위 스캔)
- 항목 메타데이터 캐시(alio_items.json) 도입 + 수동 갱신 버튼
- 다중 공시항목 × 다중 기관 일괄 다운로드 지원 (UI 트리뷰 체크박스)
- 첨부파일 다운로드 엔드포인트 일반화 (PDF/file/dfile/rule 4종 통합)
- 폴더 구조 재편: ALIO_타임스탬프 / 항목명 / 기관명 / 파일들

변경사항 (v5.2.4 → v5.3):
- 경영실적 평가결과[공기업,준정부] 공시항목 추가 (Susi API, rootNo: B1230)
- 기관별 경영실적 보고서 목록 조회 및 첨부파일 다운로드 지원

변경사항 (v5.2.3 → v5.2.4):
- SSL 인증서 검증 오류 수정 (ALIO 사이트 자체서명 인증서 대응)
- urllib3 InsecureRequestWarning 경고 메시지 비활성화
- 자체감사 결과(최근 5년) 공시항목 추가 (rootNo: 43006)

변경사항 (v5.2.2 → v5.2.3):
- 수의계약 공시항목 추가 (rootNo: 70301,70302,70303,70304)
- 분기별 수의계약 현황 조회 및 첨부파일(엑셀) 다운로드 지원

변경사항 (v5.2.1 → v5.2.2):
- 사망자수 공시항목에 첨부파일(안전경영책임보고서) 다운로드 기능 추가
- 결과내 검색 필터링 시 공시항목별 컬럼 형식 유지 버그 수정

변경사항 (v5.2.0 → v5.2.1):
- 공시항목명 변경: "산업재해 및 안전사고 사망자수" → "사망자수(최근 5년)"
- 청렴도처럼 연도별 컬럼으로 표시 (2024년~2020년)
- 산업재해+안전사고 사망자수 합산하여 단일 값으로 표시

변경사항 (v5.1.1 → v5.2.0):
- 산업재해 및 안전사고 사망자수 공시항목 추가 (rootNo: 70401)
- 산업재해 사망자수 / 안전사고 사망자수 분류 표시
- 엑셀 저장 시 사망자수 합계 컬럼 추가

변경사항 (v5.1.0 → v5.1.1):
- 징계제도 운영현황 공시항목 추가 (rootNo: 2120)

변경사항 (v5.0 → v5.1):
- 엑셀 저장 기능 개선: 공시항목별 맞춤 컬럼 적용
- 청렴도 평가 결과: 연도별 등급 컬럼 추가
- 징계처분 현황: 징계종류별 컬럼 및 합계 추가
- 환경법규 위반현황: 위반건수 컬럼 추가
- 내부규정: 규정분류별 건수 컬럼 추가
- 일반현황: 설립근거 컬럼 추가

변경사항 (v4 → v5):
- 내부규정 공시항목 추가 (정관, 인사·복무·징계, 보수, 직제, 기타)
- 내부규정 전용 API 연동 (findRuleList.json)
- 규정분류 필터 추가
"""
import atexit
import json
import os
import random
import re
import sys
import threading
import time
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from tkinter import ttk, messagebox, scrolledtext, filedialog
from typing import Callable, Optional
from urllib.parse import quote, urljoin, parse_qs, urlparse

import requests
from requests.adapters import HTTPAdapter


# ── alio_core 공유 모듈 (자기 폴더의 alio_core.py) ─────
# 본 파일은 alio-mcp 레포의 alio_core.py(정본)에서 sync된 사본.
# 직접 수정 금지 — 변경 시 alio-mcp/alio_core.py 패치 후 sync_to_crawler.sh 실행.
from alio_core import (
    sanitize_filename, create_session, retry_request,
    BASE_URL, HEADERS, INST_TYPE_CODES, DEPT_CODES, REGION_CODES,
    DISCLOSURE_ITEMS, DISCIPLINE_TYPES, RULE_DIVIS_CODES, ENDPOINT_REGISTRY,
    parse_files_field,
    fetch_alio_items, get_alio_items, build_item_display_name, build_item_root_no,
    get_alio_items_cache_path, load_alio_items_cache, save_alio_items_cache,
    group_items_by_category,
    detect_endpoint_kind, build_save_path, download_file_to_path,
    _resolve_collision_path,
    fetch_board_attachment_list, fetch_board_external_links,
    download_board_attachment, download_attachment,
    load_public_institutions,
)


# ── 공통 유틸리티 (인라인) ─────────────────────────
# 이하 코드는 shared/ 모듈에서 추출. 단독 실행을 위해 포함.

# ── Windows 네이티브 테마 적용 ────────────────────

def apply_windows_theme(root: tk.Tk) -> ttk.Style:
    style = ttk.Style(root)
    for theme in ("vista", "winnative", "xpnative"):
        if theme in style.theme_names():
            style.theme_use(theme)
            break
    return style


# ── BoundedText ──────────────────────────────────

class BoundedText(tk.Text):
    MAX_LINES = 10000

    def __init__(self, master=None, max_lines: int = MAX_LINES, **kwargs):
        super().__init__(master, **kwargs)
        self._max_lines = max_lines

    def insert(self, index, chars, *args):
        super().insert(index, chars, *args)
        self._trim_lines()

    def _trim_lines(self):
        line_count = int(self.index("end-1c").split(".")[0])
        if line_count > self._max_lines:
            overflow = line_count - self._max_lines
            self.delete("1.0", f"{overflow + 1}.0")


# ── ETACalculator ────────────────────────────────

class ETACalculator:
    def __init__(self, total: int):
        self._total = max(total, 1)
        self._start_time = time.time()

    def update(self, current: int) -> str:
        if current <= 0:
            return ""
        elapsed = time.time() - self._start_time
        rate = current / elapsed
        remaining_items = self._total - current
        if rate <= 0 or remaining_items <= 0:
            return "거의 완료"
        remaining_secs = remaining_items / rate
        return self._format_time(remaining_secs)

    @staticmethod
    def _format_time(seconds: float) -> str:
        seconds = int(seconds)
        if seconds < 5:
            return "거의 완료"
        if seconds < 60:
            return f"약 {seconds}초"
        minutes = seconds // 60
        secs = seconds % 60
        if secs == 0:
            return f"약 {minutes}분"
        return f"약 {minutes}분 {secs}초"


# ── 종료 핸들러 ──────────────────────────────────

def setup_close_handler(
    root: tk.Tk,
    cleanup_fn: Optional[Callable] = None,
    confirm_if_running: Optional[Callable[[], bool]] = None,
):
    def _on_closing():
        if confirm_if_running and confirm_if_running():
            result = messagebox.askyesno(
                "확인",
                "작업이 진행 중입니다. 프로그램을 종료하시겠습니까?"
            )
            if not result:
                return
        if cleanup_fn:
            try:
                cleanup_fn()
            except Exception:
                pass
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", _on_closing)
    if cleanup_fn:
        atexit.register(cleanup_fn)


# ── 키보드 단축키 ────────────────────────────────

def bind_shortcuts(
    root: tk.Tk,
    on_quit: Optional[Callable] = None,
    on_open: Optional[Callable] = None,
    on_stop: Optional[Callable] = None,
):
    if on_quit:
        def _quit(event=None):
            on_quit()
        root.bind("<Control-q>", _quit)
    if on_open:
        def _open(event=None):
            on_open()
        root.bind("<Control-o>", _open)
    if on_stop:
        def _stop(event=None):
            on_stop()
        root.bind("<Escape>", _stop)


# ── 설정 저장/복원 ──────────────────────────────

SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".audit_tools.json")


def _load_all() -> dict:
    if not os.path.exists(SETTINGS_FILE):
        return {}
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def _save_all(data: dict) -> None:
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except OSError:
        pass


def load_settings(app_name: str) -> dict:
    all_settings = _load_all()
    return all_settings.get(app_name, {})


def save_settings(app_name: str, settings: dict) -> None:
    all_settings = _load_all()
    all_settings[app_name] = settings
    _save_all(all_settings)


# ── 파일명 정제·네트워크 유틸은 alio_core로 이전 (v5.4.3) ──

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# HTML 파싱
from bs4 import BeautifulSoup

# 상수(BASE_URL·HEADERS·*_CODES·DISCLOSURE_ITEMS·RULE_DIVIS_CODES)는 alio_core로 이전


# parse_files_field·메뉴 API·캐시·그룹화 함수는 alio_core로 이전


# 다운로드 함수(detect_endpoint_kind·download_file_to_path·게시판형 첨부 등)는 alio_core로 이전


# load_public_institutions는 alio_core로 이전


class ItemSelectorDialog:
    """
    공시항목 다중 선택 다이얼로그 (v5.4).
    formList.json으로 가져온 알리오 전 항목을 분류 트리뷰로 표시하고
    체크박스(☑/☐)로 다중 선택. 검색·전체선택 지원.
    """
    def __init__(self, parent, items, preselected=None):
        self.items = items
        self._preselected_keys = {self._meta_key(it) for it in (preselected or [])}
        self.checked = {}  # {meta_key: BoolVar}
        self.tree_iid_to_key = {}  # 트리 iid → meta_key
        self.result = None  # 확인 시 선택된 item dict 리스트, 취소 시 None

        self.top = tk.Toplevel(parent)
        self.top.title("공시항목 선택 (v5.4)")
        self.top.geometry("760x620")
        self.top.minsize(600, 500)
        self.top.transient(parent)
        self.top.grab_set()

        self._build_ui()
        self._populate_tree()

    @staticmethod
    def _meta_key(item):
        """항목 식별 키 (mcd + scd 조합)"""
        return f"{item.get('mcd', '')}__{item.get('scd') or 'NA'}"

    def _build_ui(self):
        # 검색 + 전체선택 + 카운트
        top_frame = ttk.Frame(self.top, padding=5)
        top_frame.pack(fill=tk.X)

        ttk.Label(top_frame, text="검색:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(top_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=(2, 10))
        self.search_var.trace_add("write", lambda *a: self._populate_tree())

        ttk.Button(top_frame, text="전체선택", command=self.select_all_visible, width=9).pack(side=tk.LEFT, padx=2)
        ttk.Button(top_frame, text="전체해제", command=self.deselect_all_visible, width=9).pack(side=tk.LEFT, padx=2)
        ttk.Button(top_frame, text="모두해제", command=self.deselect_all, width=9).pack(side=tk.LEFT, padx=2)

        self.count_var = tk.StringVar(value="선택: 0개")
        ttk.Label(top_frame, textvariable=self.count_var).pack(side=tk.RIGHT)

        # 트리뷰
        tree_frame = ttk.Frame(self.top, padding=(5, 0, 5, 5))
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(
            tree_frame, columns=("rootno", "type"),
            show="tree headings", height=20
        )
        self.tree.heading("#0", text="분류 / 항목")
        self.tree.heading("rootno", text="rootNo")
        self.tree.heading("type", text="유형")
        self.tree.column("#0", width=440)
        self.tree.column("rootno", width=130, anchor="center")
        self.tree.column("type", width=80, anchor="center")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.bind("<Button-1>", self._on_click)

        # 하단 버튼
        btn_frame = ttk.Frame(self.top, padding=5)
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="확인", command=self._on_ok, width=10).pack(side=tk.RIGHT, padx=2)
        ttk.Button(btn_frame, text="취소", command=self._on_cancel, width=10).pack(side=tk.RIGHT)

    def _populate_tree(self):
        # 기존 항목 제거
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        self.tree_iid_to_key.clear()

        keyword = (self.search_var.get() or "").strip().lower()
        tree_data = group_items_by_category(self.items)

        for lcdnm, nm_dict in tree_data.items():
            l_node = self.tree.insert("", "end", text=f"📁 {lcdnm}", open=bool(keyword))
            l_count = 0
            for nmcdnm, item_list in nm_dict.items():
                # 검색 필터
                if keyword:
                    item_list = [
                        it for it in item_list
                        if keyword in build_item_display_name(it).lower()
                        or keyword in nmcdnm.lower()
                        or keyword in lcdnm.lower()
                    ]
                    if not item_list:
                        continue

                nm_node = self.tree.insert(l_node, "end", text=f"  {nmcdnm}", open=bool(keyword))
                for it in item_list:
                    key = self._meta_key(it)
                    name = build_item_display_name(it)
                    rootno = build_item_root_no(it)
                    type_label = "보고서" if (it.get("reportYn") == "Y") else "게시판"

                    # 체크 상태 - 처음 등장하면 preselected 반영
                    if key not in self.checked:
                        self.checked[key] = tk.BooleanVar(
                            value=(key in self._preselected_keys)
                        )
                    is_checked = self.checked[key].get()

                    prefix = "☑ " if is_checked else "☐ "
                    leaf_iid = self.tree.insert(
                        nm_node, "end",
                        text=prefix + name,
                        values=(rootno, type_label),
                    )
                    self.tree_iid_to_key[leaf_iid] = key
                    l_count += 1

            if l_count == 0:
                self.tree.delete(l_node)

        self._update_count()

    def _on_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        key = self.tree_iid_to_key.get(iid)
        if not key:
            return  # 분류 노드는 체크 불가
        var = self.checked.get(key)
        if not var:
            return
        var.set(not var.get())
        # 텍스트 prefix 업데이트
        cur = self.tree.item(iid, "text")
        new_prefix = "☑ " if var.get() else "☐ "
        if len(cur) >= 2 and cur[0] in ("☑", "☐"):
            cur = cur[2:]
        self.tree.item(iid, text=new_prefix + cur)
        self._update_count()

    def select_all_visible(self):
        """현재 트리에 표시된 항목만 선택"""
        for iid, key in self.tree_iid_to_key.items():
            if key in self.checked:
                self.checked[key].set(True)
        self._populate_tree()

    def deselect_all_visible(self):
        """현재 트리에 표시된 항목만 해제"""
        for iid, key in self.tree_iid_to_key.items():
            if key in self.checked:
                self.checked[key].set(False)
        self._populate_tree()

    def deselect_all(self):
        """검색 결과와 무관하게 전체 해제"""
        for v in self.checked.values():
            v.set(False)
        self._populate_tree()

    def _update_count(self):
        n = sum(1 for v in self.checked.values() if v.get())
        self.count_var.set(f"선택: {n}개")

    def _on_ok(self):
        selected_keys = {k for k, v in self.checked.items() if v.get()}
        self.result = [
            it for it in self.items if self._meta_key(it) in selected_keys
        ]
        self.top.destroy()

    def _on_cancel(self):
        self.result = None
        self.top.destroy()


class ALIOCrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("ALIO 항목별 공시 크롤링 시스템 v5.4.1")
        self.root.geometry("950x900")
        self.root.resizable(True, True)
        self.root.minsize(900, 700)

        try:
            self.root.attributes('-toolwindow', False)
        except Exception:
            pass

        self._stop_event = threading.Event()
        self.session = create_session(verify_ssl=False)
        self.session.headers.update(HEADERS)
        self.all_results = []
        self.filtered_results = []
        self.api_response_data = None
        self.public_inst_data = {}  # 공공기관 지정현황 데이터
        self.sort_column = None  # 현재 정렬 컬럼
        self.sort_reverse = False  # 정렬 방향 (False=오름차순, True=내림차순)
        self.integrity_years = []  # 청렴도 연도 목록
        # v5.4: 다중 공시항목 선택 상태
        self.selected_item_infos = []  # formList.json 메타 dict 리스트
        self.alio_items_cache = None    # 알리오 항목 캐시 (지연 로드)

        # Windows 네이티브 테마 적용
        apply_windows_theme(self.root)

        self.create_widgets()

        # 설정 복원
        APP_NAME = "alio_crawler"
        saved = load_settings(APP_NAME)
        if saved.get("last_folder"):
            self.save_path_var.set(saved["last_folder"])

        # 종료 핸들러
        setup_close_handler(
            self.root,
            cleanup_fn=self._cleanup,
            confirm_if_running=lambda: not self._stop_event.is_set() and hasattr(self, '_running_active') and self._running_active,
        )

        # 키보드 단축키
        bind_shortcuts(
            self.root,
            on_quit=self.root.destroy,
            on_open=self.browse_folder,
            on_stop=self.stop_process,
        )
    
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding=5)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ===== Title =====
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(title_frame, text="ALIO 항목별 공시 크롤링 시스템 v5.4.1",
                 font=("맑은 고딕", 11, "bold")).pack(side=tk.LEFT)
        ttk.Button(title_frame, text="도움말",
                  command=self.show_help, width=8).pack(side=tk.RIGHT)
        ttk.Button(title_frame, text="API 테스트",
                  command=self.test_api, width=10).pack(side=tk.RIGHT, padx=5)
        
        # ===== Search Options =====
        search_frame = ttk.LabelFrame(main_frame, text=" 검색 조건 ", padding=8)
        search_frame.pack(fill=tk.X, pady=(0, 5))

        # Row 1: 기관유형 + 주무부처
        row1 = ttk.Frame(search_frame)
        row1.pack(fill=tk.X, pady=2)

        ttk.Label(row1, text="기관유형:").pack(side=tk.LEFT)
        self.inst_type_var = tk.StringVar(value="전체")
        self.inst_type_combo = ttk.Combobox(
            row1, textvariable=self.inst_type_var,
            values=list(INST_TYPE_CODES.keys()),
            width=22, state="readonly"
        )
        self.inst_type_combo.pack(side=tk.LEFT, padx=(2, 15))

        ttk.Label(row1, text="주무부처:").pack(side=tk.LEFT)
        self.dept_var = tk.StringVar(value="전체")
        self.dept_combo = ttk.Combobox(
            row1, textvariable=self.dept_var,
            values=list(DEPT_CODES.keys()),
            width=22, state="readonly"
        )
        self.dept_combo.pack(side=tk.LEFT, padx=(2, 15))

        ttk.Label(row1, text="지역:").pack(side=tk.LEFT)
        self.region_var = tk.StringVar(value="전체")
        self.region_combo = ttk.Combobox(
            row1, textvariable=self.region_var,
            values=list(REGION_CODES.keys()),
            width=15, state="readonly"
        )
        self.region_combo.pack(side=tk.LEFT, padx=(2, 0))
        
        # Row 2: 기관명 + 공시항목
        row2 = ttk.Frame(search_frame)
        row2.pack(fill=tk.X, pady=5)

        ttk.Label(row2, text="기관명:").pack(side=tk.LEFT)
        self.inst_name_var = tk.StringVar(value="")
        self.inst_name_entry = ttk.Entry(row2, textvariable=self.inst_name_var, width=20)
        self.inst_name_entry.pack(side=tk.LEFT, padx=(2, 15))

        ttk.Label(row2, text="공시항목:").pack(side=tk.LEFT)
        # v5.4: 콤보박스 대신 다중 선택 버튼 + 요약 라벨
        self.item_var = tk.StringVar(value="선택하세요")  # 호환용 (첫 항목명)
        self.item_select_btn = ttk.Button(
            row2, text="공시항목 선택...",
            command=self.open_item_selector, width=16
        )
        self.item_select_btn.pack(side=tk.LEFT, padx=(2, 3))

        self.refresh_items_btn = ttk.Button(
            row2, text="갱신", command=self.refresh_alio_items, width=5
        )
        self.refresh_items_btn.pack(side=tk.LEFT, padx=(0, 8))

        self.item_summary_var = tk.StringVar(value="(선택된 항목 없음)")
        self.item_summary_label = ttk.Label(
            row2, textvariable=self.item_summary_var,
            font=("Consolas", 9), foreground="gray"
        )
        self.item_summary_label.pack(side=tk.LEFT, padx=(0, 10))

        # rootNo 표시 (호환용 stub - 다른 코드에서 참조하지만 화면 미표시)
        self.rootno_var = tk.StringVar(value="")
        self.rootno_label = ttk.Label(
            row2, textvariable=self.rootno_var,
            font=("Consolas", 9), foreground="gray"
        )
        # 화면에는 표시 안 함 (필요 시 단계 4에서 재활용)

        # 내부규정 분류 (내부규정 선택 시에만 표시)
        self.rule_divis_label = ttk.Label(row2, text="규정분류:")
        self.rule_divis_var = tk.StringVar(value="전체")
        self.rule_divis_combo = ttk.Combobox(
            row2, textvariable=self.rule_divis_var,
            values=list(RULE_DIVIS_CODES.keys()),
            width=15, state="readonly"
        )

        # v5.3 호환: item_combo는 제거되었지만 None으로 둬서 hasattr 체크 가능
        self.item_combo = None
        
        # Row 3: 저장경로
        row3 = ttk.Frame(search_frame)
        row3.pack(fill=tk.X, pady=2)

        ttk.Label(row3, text="저장경로:").pack(side=tk.LEFT)
        self.save_path_var = tk.StringVar(value=r"D:\GDRIVE\크롤링 데이터 저장소")
        self.save_path_entry = ttk.Entry(row3, textvariable=self.save_path_var)
        self.save_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 5))
        ttk.Button(row3, text="찾아보기", command=self.browse_folder, width=8).pack(side=tk.LEFT)
        
        # ===== Action Buttons + Progress =====
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=5)

        self.search_btn = ttk.Button(btn_frame, text="1.기관목록수집", command=self.start_organ_search, width=12)
        self.search_btn.pack(side=tk.LEFT, padx=(0, 3))

        self.detail_btn = ttk.Button(btn_frame, text="2.공시내용수집", command=self.start_disclosure_fetch, width=12, state=tk.DISABLED)
        self.detail_btn.pack(side=tk.LEFT, padx=(0, 3))

        self.export_btn = ttk.Button(btn_frame, text="3.엑셀저장", command=self.export_to_excel, width=10, state=tk.DISABLED)
        self.export_btn.pack(side=tk.LEFT, padx=(0, 3))

        self.stop_btn = ttk.Button(btn_frame, text="중지", command=self.stop_process, width=6, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(btn_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # ===== Filter =====
        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(fill=tk.X, pady=3)

        ttk.Label(filter_frame, text="결과내 검색:").pack(side=tk.LEFT)
        self.filter_var = tk.StringVar()
        self.filter_entry = ttk.Entry(filter_frame, textvariable=self.filter_var, width=20)
        self.filter_entry.pack(side=tk.LEFT, padx=(2, 3))
        self.filter_entry.bind('<Return>', lambda e: self.apply_filter())

        self.filter_btn = ttk.Button(filter_frame, text="적용", command=self.apply_filter, width=5, state=tk.DISABLED)
        self.filter_btn.pack(side=tk.LEFT, padx=(0, 2))
        self.clear_filter_btn = ttk.Button(filter_frame, text="해제", command=self.clear_filter, width=5, state=tk.DISABLED)
        self.clear_filter_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.select_all_btn = ttk.Button(filter_frame, text="전체선택", command=self.select_all, width=7, state=tk.DISABLED)
        self.select_all_btn.pack(side=tk.LEFT, padx=(0, 2))
        self.deselect_all_btn = ttk.Button(filter_frame, text="전체해제", command=self.deselect_all, width=7, state=tk.DISABLED)
        self.deselect_all_btn.pack(side=tk.LEFT)

        self.result_count_var = tk.StringVar(value="")
        ttk.Label(filter_frame, textvariable=self.result_count_var).pack(side=tk.RIGHT)
        
        # ===== Results Treeview =====
        result_frame = ttk.LabelFrame(main_frame, text=" 검색 결과 (기관 목록) ", padding=3)
        result_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 3))

        # 13개 열 (징계처분 현황용 확장)
        columns = ("select", "no", "apba_id", "inst_name", "inst_type", "dept", "col6", "col7", "col8", "col9", "col10", "col11", "col12")
        self.result_tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=12)

        self.result_tree.heading("select", text="선택")
        self.result_tree.heading("no", text="번호", command=lambda: self.sort_treeview("no"))
        self.result_tree.heading("apba_id", text="기관코드", command=lambda: self.sort_treeview("apba_id"))
        self.result_tree.heading("inst_name", text="기관명", command=lambda: self.sort_treeview("inst_name"))
        self.result_tree.heading("inst_type", text="기관유형", command=lambda: self.sort_treeview("inst_type"))
        self.result_tree.heading("dept", text="주무부처", command=lambda: self.sort_treeview("dept"))
        self.result_tree.heading("col6", text="지역", command=lambda: self.sort_treeview("col6"))
        self.result_tree.heading("col7", text="공시기간", command=lambda: self.sort_treeview("col7"))
        self.result_tree.heading("col8", text="첨부파일", command=lambda: self.sort_treeview("col8"))
        self.result_tree.heading("col9", text="공시번호", command=lambda: self.sort_treeview("col9"))
        self.result_tree.heading("col10", text="", command=lambda: self.sort_treeview("col10"))
        self.result_tree.heading("col11", text="", command=lambda: self.sort_treeview("col11"))
        self.result_tree.heading("col12", text="", command=lambda: self.sort_treeview("col12"))

        self.result_tree.column("select", width=40, anchor="center")
        self.result_tree.column("no", width=45, anchor="center")
        self.result_tree.column("apba_id", width=70, anchor="center")
        self.result_tree.column("inst_name", width=180)
        self.result_tree.column("inst_type", width=130)
        self.result_tree.column("dept", width=120)
        self.result_tree.column("col6", width=90)
        self.result_tree.column("col7", width=90, anchor="center")
        self.result_tree.column("col8", width=60, anchor="center")
        self.result_tree.column("col9", width=130, anchor="center")
        self.result_tree.column("col10", width=0)
        self.result_tree.column("col11", width=0)
        self.result_tree.column("col12", width=0)
        
        vsb = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.result_tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient=tk.HORIZONTAL, command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.result_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        result_frame.grid_rowconfigure(0, weight=1)
        result_frame.grid_columnconfigure(0, weight=1)
        
        self.result_tree.bind('<ButtonRelease-1>', self.on_tree_click)
        self.result_tree.bind('<Double-1>', self.on_tree_double_click)
        
        # ===== Log =====
        log_frame = ttk.LabelFrame(main_frame, text=" 진행 상황 ", padding=3)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 3))

        # 로그 버튼 프레임
        log_btn_frame = ttk.Frame(log_frame)
        log_btn_frame.pack(fill=tk.X, pady=(0, 3))
        ttk.Button(log_btn_frame, text="전체선택", command=self.select_all_log, width=8).pack(side=tk.LEFT, padx=(0, 3))
        ttk.Button(log_btn_frame, text="복사", command=self.copy_log, width=6).pack(side=tk.LEFT)

        # BoundedText (줄 수 제한 tk.Text)
        log_scroll_frame = ttk.Frame(log_frame)
        log_scroll_frame.pack(fill=tk.BOTH, expand=True)
        self.log_text = BoundedText(log_scroll_frame, height=10, wrap=tk.WORD,
                                     font=("Consolas", 9))
        log_scrollbar = ttk.Scrollbar(log_scroll_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # ===== Status Bar =====
        self.status_var = tk.StringVar(value="준비")
        ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W).pack(fill=tk.X)
        ttk.Label(main_frame, text="문의: giovinazo@yahoo.co.kr", anchor=tk.E,
                 foreground="gray", font=("TkDefaultFont", 9)).pack(fill=tk.X, padx=4)

        self.selected_items = set()
    
    def show_help(self):
        help_text = (
            "【 ALIO 항목별 공시 크롤러 v5.4.1 사용법 】\n\n"
            "1. 기관목록수집: 선택한 조건에 맞는 기관 목록을 가져옵니다.\n"
            "   - 기관유형: 공기업(시장형/준시장형), 준정부기관(기금관리형/위탁집행형), 기타공공기관\n"
            "   - 주무부처: 44개 부처/청 중 선택\n"
            "   - 지역: 17개 시도 중 선택\n\n"
            "2. 공시내용수집: 선택한 기관들의 공시 상세 내용을 수집합니다.\n"
            "   - PDF 파일 다운로드\n"
            "   - 첨부파일 다운로드\n\n"
            "3. 엑셀저장: 수집된 데이터를 엑셀 파일로 저장합니다.\n\n"
            "【 필터 조건 】\n"
            "- 기관유형 5개: 공기업(시장형), 공기업(준시장형), 준정부기관(기금관리형), 준정부기관(위탁집행형), 기타공공기관\n"
            "- 주무부처 44개: 2025년 1월 기준 최신화\n"
            "- 지역 17개: 전국 시도\n\n"
            "【 주의사항 】\n"
            "- 더블클릭: 해당 기관의 공시 페이지를 브라우저로 열기\n"
            "- 과도한 요청은 서버에 부담을 줄 수 있으니 주의하세요."
        )
        messagebox.showinfo("도움말", help_text)

    def select_all_log(self, event=None):
        """로그 텍스트 전체선택"""
        self.log_text.tag_add('sel', '1.0', 'end')
        self.log_text.focus_set()
        return 'break'

    def copy_log(self, event=None):
        """선택된 로그 복사 (선택 없으면 전체 복사)"""
        try:
            text = self.log_text.get('sel.first', 'sel.last')
        except tk.TclError:
            text = self.log_text.get('1.0', 'end-1c')
        self.root.clipboard_clear()
        self.root.clipboard_append(text)

    def test_api(self):
        """API 연결 테스트"""
        self.log_text.delete(1.0, tk.END)
        self.log("=" * 60)
        self.log("ALIO API 연결 테스트 시작")
        self.log("=" * 60)
        
        thread = threading.Thread(target=self._run_api_test)
        thread.daemon = True
        thread.start()
    
    def _run_api_test(self):
        """API 테스트 실행"""
        item_info = self.get_selected_item_info()
        root_no = item_info["rootNo"] if item_info else "20801"
        
        self.log("\n[1] itemOrganListJung.json API 테스트")
        api_url = f"{BASE_URL}/item/itemOrganListJung.json"
        
        try:
            params = {"reportFormRootNo": root_no}
            self.log(f"    URL: {api_url}")
            self.log(f"    파라미터: {params}")
            
            resp = retry_request(self.session, "POST", api_url, data=params, timeout=30)
            self.log(f"    HTTP 상태: {resp.status_code}")
            
            if resp.status_code == 200:
                try:
                    data = resp.json()
                    self.log(f"    ✓ JSON 응답 수신")
                    
                    if isinstance(data, dict):
                        self.log(f"    응답 키: {list(data.keys())}")
                        
                        if 'data' in data and data['data']:
                            inner = data['data']
                            self.log(f"    data 내부 키: {list(inner.keys())}")
                            
                            organ_list = inner.get('organList', [])
                            total_cnt = inner.get('totalCnt', 0)
                            
                            self.log(f"    organList: {len(organ_list)}개")
                            self.log(f"    totalCnt: {total_cnt}")
                            
                            if organ_list:
                                first = organ_list[0]
                                self.log(f"\n    첫 번째 항목 필드:")
                                for key in first.keys():
                                    val = first[key]
                                    val_str = str(val)[:50] if val else "(없음)"
                                    self.log(f"      - {key}: {val_str}")
                                    
                except json.JSONDecodeError as e:
                    self.log(f"    ✗ JSON 파싱 실패: {e}")
            else:
                self.log(f"    ✗ HTTP 오류: {resp.status_code}")
                
        except Exception as e:
            self.log(f"    ✗ 오류: {str(e)}")
        
        self.log("\n" + "=" * 60)
        self.log("테스트 완료")
        self.log("=" * 60)
    
    def browse_folder(self):
        folder = filedialog.askdirectory(initialdir=self.save_path_var.get() or os.getcwd())
        if folder:
            self.save_path_var.set(folder)
            save_settings("alio_crawler", {"last_folder": folder})
    
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    # ── v5.4: 다중 공시항목 선택 ────────────────
    def open_item_selector(self):
        """공시항목 다중 선택 다이얼로그 열기"""
        if self.alio_items_cache is None:
            self.alio_items_cache = get_alio_items()
            if not self.alio_items_cache:
                messagebox.showerror(
                    "오류",
                    "알리오 항목 메뉴를 불러오지 못했습니다.\n갱신 버튼을 눌러 재시도하세요."
                )
                return

        dlg = ItemSelectorDialog(
            self.root,
            items=self.alio_items_cache,
            preselected=self.selected_item_infos,
        )
        self.root.wait_window(dlg.top)

        if dlg.result is not None:
            self.selected_item_infos = dlg.result
            self._refresh_item_summary()
            # 호환: item_var에 첫 번째 항목명 설정
            if self.selected_item_infos:
                self.item_var.set(build_item_display_name(self.selected_item_infos[0]))
            else:
                self.item_var.set("선택하세요")
            # 내부규정 선택 시 규정분류 활성화
            self._toggle_rule_divis()
            self.on_items_changed()

    def refresh_alio_items(self):
        """항목 캐시 강제 갱신 (formList.json 재호출)"""
        threading.Thread(
            target=self._refresh_alio_items_worker,
            daemon=True
        ).start()

    def _refresh_alio_items_worker(self):
        def progress(cur, tot, msg):
            self.log(msg)
        items = get_alio_items(force_refresh=True, progress_callback=progress)
        if items:
            self.alio_items_cache = items
            self.root.after(
                0,
                lambda: messagebox.showinfo("갱신 완료", f"항목 {len(items)}개 갱신 완료")
            )
        else:
            self.root.after(
                0,
                lambda: messagebox.showerror("오류", "항목 갱신 실패")
            )

    def _refresh_item_summary(self):
        """선택된 항목 요약 라벨 업데이트"""
        n = len(self.selected_item_infos)
        if n == 0:
            self.item_summary_var.set("(선택된 항목 없음)")
        elif n == 1:
            name = build_item_display_name(self.selected_item_infos[0])
            self.item_summary_var.set(f"선택: {name}")
        else:
            first = build_item_display_name(self.selected_item_infos[0])
            self.item_summary_var.set(f"선택: {first} 외 {n-1}개")

    def _toggle_rule_divis(self):
        """내부규정 포함 시 규정분류 콤보 표시"""
        has_rule = any(
            (it.get("mcd") == "21110") for it in self.selected_item_infos
        )
        if has_rule:
            self.rule_divis_label.pack(side=tk.LEFT, padx=(10, 2))
            self.rule_divis_combo.pack(side=tk.LEFT, padx=(0, 10))
        else:
            self.rule_divis_label.pack_forget()
            self.rule_divis_combo.pack_forget()

    def on_items_changed(self):
        """다중 선택 변경 시 후속 처리. v5.3 on_item_changed의 신 호환판."""
        infos = self.get_selected_item_infos()
        if not infos:
            self.rootno_var.set("")
            return
        # rootNo 표시 (호환용)
        if len(infos) == 1:
            self.rootno_var.set(f"rootNo: {infos[0]['rootNo']}")
        else:
            self.rootno_var.set(f"rootNo: 다중({len(infos)}개)")

    # ── v5.3 호환 ──────────────────────────────
    def on_item_changed(self, event=None):
        """v5.3 콤보박스 시절 호환 stub. v5.4 다중 선택은 open_item_selector 사용."""
        self.on_items_changed()

    def _item_meta_to_legacy(self, item_meta):
        """
        formList.json 메타 → v5.3 DISCLOSURE_ITEMS 호환 dict.
        다단계 매칭으로 v5.3 등록 11개 항목 최대한 보존.
        """
        if "type" in item_meta and "rootNo" in item_meta and "_meta" not in item_meta:
            return item_meta  # 이미 legacy 형식

        root_no = build_item_root_no(item_meta)
        name = build_item_display_name(item_meta)
        mcd = item_meta.get("mcd", "") or ""
        kind = detect_endpoint_kind(item_meta)

        # 1) rootNo 완전 일치
        for legacy_name, legacy_info in DISCLOSURE_ITEMS.items():
            if legacy_info["rootNo"] and legacy_info["rootNo"] == root_no:
                return {**legacy_info, "name": legacy_name,
                        "_meta": item_meta, "_kind": kind}

        # 2) 내부규정 특수 매칭 (v5.3 rootNo="" but formList mcd=21110)
        if mcd == "21110":
            rule_info = DISCLOSURE_ITEMS.get("내부규정")
            if rule_info:
                return {**rule_info, "name": "내부규정",
                        "_meta": item_meta, "_kind": "rule"}

        # 3) 다중 rootNos 집합 일치 (예: v5.3 "21214,21211,21212,21213" vs formList "21211,21212,21213,21214")
        new_set = {rn.strip() for rn in root_no.split(",") if rn.strip()}
        for legacy_name, legacy_info in DISCLOSURE_ITEMS.items():
            legacy_set = {rn.strip() for rn in (legacy_info["rootNo"] or "").split(",") if rn.strip()}
            if legacy_set and legacy_set == new_set:
                return {**legacy_info, "name": legacy_name,
                        "_meta": item_meta, "_kind": kind}

        # 4) 항목명 정확 일치 (v5.3 항목명이 formList scdnm/mcdnm에 있는 경우)
        for legacy_name, legacy_info in DISCLOSURE_ITEMS.items():
            # 괄호 내용 제거 후 비교
            legacy_base = legacy_name.split("(")[0].strip()
            scdnm = (item_meta.get("scdnm") or "").strip()
            mcdnm = (item_meta.get("mcdnm") or "").strip()
            if legacy_base and (legacy_base == scdnm or legacy_base == mcdnm):
                return {**legacy_info, "name": legacy_name,
                        "_meta": item_meta, "_kind": kind}

        # 5) 신규 항목 type 결정 (v5.4.1: 게시판형 → envlaw 흐름 재활용)
        if kind == "rule":
            v53_type = "rule"
        elif (item_meta.get("reportYn") or "").upper() == "N":
            # 게시판형(reportYn=N)은 envlaw 처리 흐름과 호환 가능
            # (itemReportListSusi.json + 첨부파일 다운로드)
            v53_type = "envlaw"
        else:
            v53_type = "jung"

        return {
            "rootNo": root_no,
            "type": v53_type,
            "name": name,
            "_meta": item_meta,
            "_kind": kind,
        }

    def get_selected_item_info(self):
        """v5.3 호환: 단일 항목 dict 반환 (다중 선택 시 첫 번째)."""
        if self.selected_item_infos:
            return self._item_meta_to_legacy(self.selected_item_infos[0])
        # v5.3 콤보박스 호환 fallback
        item_name = self.item_var.get()
        if item_name in DISCLOSURE_ITEMS:
            return DISCLOSURE_ITEMS[item_name]
        return None

    def get_selected_item_infos(self):
        """v5.4: 선택된 모든 항목의 legacy 형식 리스트 반환"""
        return [self._item_meta_to_legacy(m) for m in self.selected_item_infos]
    
    def on_tree_click(self, event):
        region = self.result_tree.identify("region", event.x, event.y)
        if region == "cell":
            item = self.result_tree.identify_row(event.y)
            if item:
                if item in self.selected_items:
                    self.selected_items.discard(item)
                    values = list(self.result_tree.item(item, "values"))
                    values[0] = "[ ]"
                    self.result_tree.item(item, values=values)
                else:
                    self.selected_items.add(item)
                    values = list(self.result_tree.item(item, "values"))
                    values[0] = "[v]"
                    self.result_tree.item(item, values=values)
                self.update_selection_count()
    
    def on_tree_double_click(self, event):
        item = self.result_tree.identify_row(event.y)
        if item:
            values = self.result_tree.item(item, "values")
            apba_id = values[2]
            item_info = self.get_selected_item_info()
            if item_info and apba_id:
                url = f"{BASE_URL}/item/itemReportTerm.do?apbaId={apba_id}&reportFormRootNo={item_info['rootNo']}"
                import webbrowser
                webbrowser.open(url)

    def sort_treeview(self, col):
        """트리뷰 컬럼 정렬 (오름차순/내림차순 토글)"""
        # 같은 컬럼 클릭 시 정렬 방향 토글
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False

        # 현재 데이터 가져오기
        items = [(self.result_tree.item(item, "values"), item) for item in self.result_tree.get_children()]
        if not items:
            return

        # 컬럼 인덱스 찾기
        columns = ("select", "no", "apba_id", "inst_name", "inst_type", "dept", "col6", "col7", "col8", "col9", "col10", "col11", "col12")
        col_idx = columns.index(col) if col in columns else 0

        # 정렬 (숫자는 숫자로, 문자는 문자로)
        def sort_key(item):
            val = item[0][col_idx] if col_idx < len(item[0]) else ""
            # 숫자 변환 시도
            try:
                # "4개" 같은 형식 처리
                if isinstance(val, str) and val.endswith("개"):
                    return (0, int(val[:-1]))
                return (0, int(val))
            except (ValueError, TypeError):
                return (1, str(val).lower())

        items.sort(key=sort_key, reverse=self.sort_reverse)

        # 정렬된 순서로 재배치
        for idx, (values, item) in enumerate(items):
            self.result_tree.move(item, '', idx)

        # 헤더에 정렬 방향 표시 업데이트
        for c in columns:
            current_text = self.result_tree.heading(c, "text")
            # 기존 화살표 제거
            current_text = current_text.replace(" ▲", "").replace(" ▼", "")
            if c == col:
                arrow = " ▼" if self.sort_reverse else " ▲"
                self.result_tree.heading(c, text=current_text + arrow)
            else:
                self.result_tree.heading(c, text=current_text)

    def select_all(self):
        for item in self.result_tree.get_children():
            self.selected_items.add(item)
            values = list(self.result_tree.item(item, "values"))
            values[0] = "[v]"
            self.result_tree.item(item, values=values)
        self.update_selection_count()
    
    def deselect_all(self):
        for item in self.result_tree.get_children():
            self.selected_items.discard(item)
            values = list(self.result_tree.item(item, "values"))
            values[0] = "[ ]"
            self.result_tree.item(item, values=values)
        self.update_selection_count()
    
    def update_selection_count(self):
        total = len(self.result_tree.get_children())
        selected = len(self.selected_items)

        # 내부규정일 때 규정 수 요약 추가
        item_info = self.get_selected_item_info()
        if item_info and item_info.get("type") == "rule" and self.filtered_results:
            total_rules = sum(r.get("rule_total", 0) for r in self.filtered_results)
            # 규정분류별 집계
            all_divis = {}
            for r in self.filtered_results:
                for k, v in r.get("rule_by_divis", {}).items():
                    all_divis[k] = all_divis.get(k, 0) + v
            divis_str = ", ".join([f"{k}({v})" for k, v in sorted(all_divis.items(), key=lambda x: -x[1])[:5]])
            self.result_count_var.set(f"기관: {total}개 / 선택: {selected}개 | 총 규정: {total_rules}건 [{divis_str}]")
        else:
            self.result_count_var.set(f"전체: {total}건 / 선택: {selected}건")
    
    def apply_filter(self):
        keyword = self.filter_var.get().strip().lower()
        
        children = self.result_tree.get_children()
        if children:
            self.result_tree.delete(*children)
        
        self.filtered_results = []
        for case in self.all_results:
            if not keyword:
                self.filtered_results.append(case)
            else:
                searchable = f"{case.get('inst_name', '')} {case.get('inst_type', '')} {case.get('dept', '')} {case.get('region', '')}".lower()
                if keyword in searchable:
                    self.filtered_results.append(case)
        
        # 현재 공시항목 타입 확인
        item_info = self.get_selected_item_info()
        item_type = item_info.get("type", "jung") if item_info else "jung"
        
        is_rule = item_type == "rule"
        is_discipline = item_type == "discipline"
        is_integrity = item_type == "integrity"
        is_safety = item_type == "safety"
        is_envlaw = item_type == "envlaw"
        is_general = item_type == "general"
        is_audit = item_type == "audit"
        is_mgmt_eval = item_type == "mgmt_eval"

        self.selected_items.clear()
        for idx, case in enumerate(self.filtered_results, 1):
            inst_type = case.get("inst_type", "")
            dept = case.get("dept", "")
            
            if is_general:
                purpose = case.get("purpose", "-")
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    purpose[:50] + "..." if len(purpose) > 50 else purpose,
                    "", "", "", "", "", ""
                )
            elif is_discipline:
                dc = case.get("discipline_counts", {})
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    dc.get("파면", 0), dc.get("해임", 0), dc.get("정직", 0),
                    dc.get("감봉", 0), dc.get("견책", 0), dc.get("기타", 0), ""
                )
            elif is_integrity:
                grades = case.get("integrity_grades", {})
                years = getattr(self, 'integrity_years', [str(datetime.now().year - i) for i in range(5)])
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    grades.get(years[0], "-") if len(years) > 0 else "-",
                    grades.get(years[1], "-") if len(years) > 1 else "-",
                    grades.get(years[2], "-") if len(years) > 2 else "-",
                    grades.get(years[3], "-") if len(years) > 3 else "-",
                    grades.get(years[4], "-") if len(years) > 4 else "-",
                    "", ""
                )
            elif is_safety:
                death_by_year = case.get("death_by_year", {})
                years = getattr(self, 'safety_years', [str(datetime.now().year - i) for i in range(5)])
                y0 = death_by_year.get(years[0], 0) if len(years) > 0 else 0
                y1 = death_by_year.get(years[1], 0) if len(years) > 1 else 0
                y2 = death_by_year.get(years[2], 0) if len(years) > 2 else 0
                y3 = death_by_year.get(years[3], 0) if len(years) > 3 else 0
                y4 = death_by_year.get(years[4], 0) if len(years) > 4 else 0
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    f"{y0}명" if y0 > 0 else "0",
                    f"{y1}명" if y1 > 0 else "0",
                    f"{y2}명" if y2 > 0 else "0",
                    f"{y3}명" if y3 > 0 else "0",
                    f"{y4}명" if y4 > 0 else "0",
                    "", ""
                )
            elif is_envlaw:
                violation_count = case.get("violation_count", 0)
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    f"{violation_count}건" if violation_count > 0 else "0",
                    "", "", "", "", "", ""
                )
            elif is_audit:
                audit_count = case.get("audit_count", 0)
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    f"{audit_count}건" if audit_count > 0 else "0",
                    "", "", "", "", "", ""
                )
            elif is_mgmt_eval:
                report_count = case.get("report_count", 0)
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    case.get("region", "") or "-",
                    f"{report_count}건" if report_count > 0 else "0",
                    case.get("latest_title", "-")[:30],
                    case.get("latest_idate", "-"),
                    "", "", ""
                )
            elif is_rule:
                rule_total = case.get("rule_total", 0)
                rule_by_divis = case.get("rule_by_divis", {})
                if rule_by_divis:
                    divis_parts = [f"{k}({v})" for k, v in rule_by_divis.items()]
                    divis_str = ", ".join(divis_parts[:4])
                    if len(divis_parts) > 4:
                        divis_str += "..."
                else:
                    divis_str = "-"
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    case.get("region", "") or "-",
                    f"{rule_total}건" if rule_total else "0건",
                    divis_str, "", "", "", ""
                )
            else:
                # 기본 타입 (jung 등)
                files_count = len(case.get("files_parsed", []))
                values = (
                    "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                    inst_type, dept,
                    case.get("region", "") or "-",
                    case.get("period", "") or "-",
                    f"{files_count}개" if files_count else "-",
                    case.get("disclosure_no", "")[:20] if case.get("disclosure_no") else "-",
                    "", "", ""
                )
            
            item_id = self.result_tree.insert("", tk.END, values=values)
            case["tree_id"] = item_id
        
        self.update_selection_count()
        
        if keyword:
            self.log(f"필터 적용: '{keyword}' → {len(self.filtered_results)}건")
    
    def clear_filter(self):
        self.filter_var.set("")
        self.apply_filter()
    
    def start_organ_search(self):
        self._stop_event.clear()
        self._running_active = True
        self.search_btn.config(state=tk.DISABLED)
        self.detail_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress_var.set(0)
        self.log_text.delete(1.0, tk.END)
        
        children = self.result_tree.get_children()
        if children:
            self.result_tree.delete(*children)
        self.all_results = []
        self.filtered_results = []
        self.selected_items.clear()
        
        thread = threading.Thread(target=self.search_organs)
        thread.daemon = True
        thread.start()
    
    def search_organs(self):
        """기관 목록 검색"""
        try:
            item_info = self.get_selected_item_info()
            item_name = self.item_var.get()
            inst_type_filter = self.inst_type_var.get()
            dept_filter = self.dept_var.get()
            region_filter = self.region_var.get()
            inst_name_filter = self.inst_name_var.get().strip()

            if not item_info:
                self.log("오류: 공시항목을 선택해 주세요.")
                self.finish_process(False)
                return

            root_no = item_info["rootNo"]

            self.log("=" * 60)
            self.log("ALIO 기관 목록 수집 시작 (v5.4.1)")
            self.log("=" * 60)
            self.log(f"공시항목: {item_name} (rootNo: {root_no})")
            self.log(f"기관유형 필터: {inst_type_filter}")
            self.log(f"주무부처 필터: {dept_filter}")
            self.log(f"지역 필터: {region_filter}")
            if inst_name_filter:
                self.log(f"기관명 필터: {inst_name_filter}")
            self.log("")

            item_type = item_info.get("type", "jung")

            organ_list = []

            # 모든 공시항목에서 지역 필터링을 위해 기관목록 API 로드
            if region_filter != "전체" or not self.public_inst_data:
                self.log("ALIO 기관목록 API 로드 중 (지역 정보)...")

                def on_progress(current, total):
                    pct = current / total * 30  # 30%까지 진행
                    self.progress_var.set(pct)
                    self.status_var.set(f"기관목록 로드 중... ({current}/{total} 페이지)")

                self.public_inst_data = load_public_institutions(progress_callback=on_progress)
                self.log(f"공공기관 {len(self.public_inst_data)}개 로드 완료")
                self.log("")

            # 내부규정: 공공기관 목록 기반
            if item_type == "rule":
                rule_divis = RULE_DIVIS_CODES.get(self.rule_divis_var.get(), "")
                rule_divis_name = self.rule_divis_var.get()
                self.log(f"규정분류: {rule_divis_name}")
                self.log("")

                # 기관목록이 아직 로드되지 않은 경우에만 로드
                if not self.public_inst_data:
                    self.log("ALIO 기관목록 API 로드 중...")

                    def on_progress(current, total):
                        pct = current / total * 30
                        self.progress_var.set(pct)
                        self.status_var.set(f"기관목록 로드 중... ({current}/{total} 페이지)")

                    self.public_inst_data = load_public_institutions(progress_callback=on_progress)
                    self.log(f"공공기관 {len(self.public_inst_data)}개 로드 완료")
                    self.log("")

                # 필터 적용하여 기관 목록 생성
                filtered_list = []
                for inst_name, inst_info in self.public_inst_data.items():
                    if self._stop_event.is_set():
                        return

                    matched_inst_type = inst_info.get("inst_type", "")
                    matched_dept = inst_info.get("dept", "")
                    matched_region = inst_info.get("region", "")

                    # 기관유형 필터
                    if inst_type_filter != "전체" and inst_type_filter not in matched_inst_type:
                        continue

                    # 주무부처 필터
                    if dept_filter != "전체" and dept_filter not in matched_dept:
                        continue

                    # 지역 필터
                    if region_filter != "전체" and region_filter not in matched_region:
                        continue

                    # 기관명 필터
                    if inst_name_filter and inst_name_filter.lower() not in inst_name.lower():
                        continue

                    filtered_list.append({
                        "apba_id": inst_info.get("apba_id", ""),
                        "inst_name": inst_name,
                        "inst_type": matched_inst_type,
                        "dept": matched_dept,
                        "region": matched_region,
                    })

                self.log(f"필터 적용 후: {len(filtered_list)}개 기관")
                self.log("")
                self.log("각 기관별 규정 개수 조회 중... (병렬 처리)")

                # 규정 개수 조회 함수 (재시도 포함) - 항상 전체 분류 조회
                def fetch_rule_count(inst_info):
                    inst_name = inst_info["inst_name"]
                    rule_total = 0
                    rule_by_divis = {}

                    rule_url = f"{BASE_URL}/occasional/findRuleList.json"

                    # 항상 모든 분류별로 API 호출하여 총 규정 개수 조회
                    divis_codes = {
                        "정관": "K1500",
                        "인사·복무·징계": "K1100",
                        "보수": "K1200",
                        "직제": "K1300",
                        "기타": "K1400",
                    }
                    for divis_name, divis_code in divis_codes.items():
                        params = {
                            "type": "apbaNa",
                            "word": inst_name,
                            "pageNo": 1,
                            "divis": divis_code
                        }
                        for attempt in range(3):
                            try:
                                resp = retry_request(self.session, "GET", rule_url, params=params, timeout=15)
                                if resp.status_code == 200:
                                    data = resp.json()
                                    cnt = data.get("data", {}).get("totalCnt", 0)
                                    if cnt > 0:
                                        rule_by_divis[divis_name] = cnt
                                        rule_total += cnt
                                    break
                            except requests.RequestException:
                                if attempt < 2:
                                    time.sleep(0.2)
                                continue

                    return {
                        "apba_id": inst_info["apba_id"],
                        "inst_name": inst_name,
                        "inst_type": inst_info["inst_type"],
                        "dept": inst_info["dept"],
                        "region": inst_info["region"],
                        "rule_divis": rule_divis,
                        "rule_total": rule_total,
                        "rule_by_divis": rule_by_divis,
                        "period": "",
                        "disclosure_no": "",
                        "files_parsed": [],
                        "raw_data": {},
                    }

                # 병렬 처리 (최대 10개 스레드)
                completed = 0
                eta_calc = ETACalculator(len(filtered_list))
                with ThreadPoolExecutor(max_workers=5) as executor:
                    futures = {executor.submit(fetch_rule_count, inst): inst for inst in filtered_list}

                    for future in as_completed(futures):
                        if self._stop_event.is_set():
                            return

                        result = future.result()
                        organ_list.append(result)
                        completed += 1

                        # 진행상황 표시
                        if completed % 10 == 0 or completed == len(filtered_list):
                            progress = completed / len(filtered_list) * 100
                            self.progress_var.set(progress)
                            eta_str = eta_calc.update(completed)
                            self.status_var.set(f"규정 개수 조회 중... ({completed}/{len(filtered_list)}) - {eta_str}")

                # 규정 개수 기준 정렬 (많은 순)
                organ_list.sort(key=lambda x: x.get("rule_total", 0), reverse=True)

                total_rules = sum(r.get("rule_total", 0) for r in organ_list)
                has_rules = len([r for r in organ_list if r.get("rule_total", 0) > 0])
                self.log(f"규정 보유 기관: {has_rules}개 / 총 규정: {total_rules}건")

            elif item_type == "discipline":
                # 징계처분 현황: 기관별 누적 집계
                self.log("징계처분 현황 API 호출 중...")
                api_url = f"{BASE_URL}/item/itemOrganListJung.json"

                request_body = {
                    "apbaId": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "quart": "",
                    "reportFormRootNo": root_no
                }

                headers = {
                    "Content-Type": "application/json;charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Origin": "https://www.alio.go.kr",
                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                }

                try:
                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        # 알리오 서버 응답 status 검증 (일부 항목은 서버에서 'fail' 반환)
                        if isinstance(data, dict) and data.get("status") and data.get("status") != "success":
                            srv_msg = data.get("message", "알 수 없음")
                            self.log(f"⚠ 알리오 서버 응답 실패: {srv_msg}")
                            self.log(f"  해당 공시항목(rootNo={root_no})은 알리오 사이트 자체 결함일 수 있습니다.")
                            self.log(f"  다른 공시항목을 시도하거나 잠시 후 다시 시도해 주세요.")
                            return

                        items = []
                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []
                                elif isinstance(inner_data, list):
                                    items = inner_data
                            elif 'organList' in data:
                                items = data.get('organList', []) or []

                        self.log(f"API 응답: {len(items)}건")

                        # 기관별로 그룹화
                        organ_grouped = {}

                        for item in items:
                            if self._stop_event.is_set():
                                return

                            apba_id = item.get("apbaId", "")
                            apba_na = item.get("apbaNa", "")
                            type_na = item.get("typeNa", "")
                            jidt_na = item.get("jidtNa", "")

                            # 필터링
                            if inst_type_filter != "전체" and inst_type_filter not in type_na:
                                continue
                            if dept_filter != "전체" and dept_filter not in jidt_na:
                                continue

                            # 지역 정보
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")

                            if region_filter != "전체" and region_filter not in area_na:
                                continue

                            if inst_name_filter and inst_name_filter.lower() not in apba_na.lower():
                                continue

                            # 기관별 그룹화
                            if apba_id not in organ_grouped:
                                organ_grouped[apba_id] = {
                                    "apba_id": apba_id,
                                    "inst_name": apba_na,
                                    "inst_type": type_na,
                                    "dept": jidt_na,
                                    "region": area_na,
                                    "discipline_counts": {dt: 0 for dt in DISCIPLINE_TYPES},
                                    "files_parsed": [],
                                    "disclosure_nos": [],
                                    "quarters": [],
                                    "raw_items": []
                                }

                            org = organ_grouped[apba_id]
                            org["raw_items"].append(item)

                            # 분기 정보
                            crit_yyyy = item.get("critYyyy", "")
                            quart_na = item.get("quartNa", "")
                            period = f"{crit_yyyy} {quart_na}".strip()
                            if period and period not in org["quarters"]:
                                org["quarters"].append(period)

                            # 공시번호
                            disclosure_no = item.get("disclosureNo", "")
                            if disclosure_no and disclosure_no not in org["disclosure_nos"]:
                                org["disclosure_nos"].append(disclosure_no)

                            # 첨부파일 (중복 제거)
                            files_str = item.get("files", "")
                            if files_str:
                                files_parsed = parse_files_field(files_str)
                                existing_ids = {f.get("id") for f in org["files_parsed"]}
                                for fp in files_parsed:
                                    if fp.get("id") not in existing_ids:
                                        fp["period"] = period
                                        fp["disclosure_no"] = disclosure_no
                                        org["files_parsed"].append(fp)
                                        existing_ids.add(fp.get("id"))

                            # 징계 종류별 집계 (API 응답에서 추출)
                            # 일반적으로 징계처분 데이터는 상세 API에서 가져와야 함
                            # 여기서는 기본 구조만 설정

                        # 그룹화된 데이터를 리스트로 변환
                        for apba_id, org in organ_grouped.items():
                            organ_list.append(org)

                        self.log(f"기관별 그룹화: {len(organ_list)}개 기관")

                        # 징계 상세 데이터 수집 (병렬 처리)
                        if organ_list:
                            self.log("징계처분 상세 데이터 수집 중...")

                            # 기타로 분류되는 징계종류 수집용
                            other_types_found = {}

                            def fetch_discipline_detail(org):
                                """기관별 징계처분 상세 조회 - doc.html 파싱"""
                                discipline_counts = {dt: 0 for dt in DISCIPLINE_TYPES}
                                org_other_types = []

                                for raw_item in org.get("raw_items", []):
                                    disclosure_no = raw_item.get("disclosureNo", "")
                                    if not disclosure_no:
                                        continue

                                    try:
                                        # disclosureNo에서 날짜 추출 (앞 8자리: YYYYMMDD)
                                        date_str = disclosure_no[:8]
                                        year = date_str[:4]
                                        month = date_str[4:6]
                                        day = date_str[6:8]

                                        # doc.html URL 생성
                                        doc_url = f"https://www.alio.go.kr/upload/disclosure/{year}/{month}/{day}/{disclosure_no}/doc.html"
                                        doc_resp = retry_request(self.session, "GET", doc_url, timeout=15)

                                        if doc_resp.status_code == 200:
                                            soup = BeautifulSoup(doc_resp.text, 'html.parser')

                                            # 징계처분 결과 테이블 찾기
                                            target_table = None
                                            for table in soup.find_all('table', border='1'):
                                                th = table.find('th')
                                                if th and '징계처분일' in th.get_text():
                                                    target_table = table
                                                    break

                                            if target_table:
                                                tbody = target_table.find('tbody')
                                                if tbody:
                                                    rows = tbody.find_all('tr')
                                                    for row in rows:
                                                        cells = row.find_all('td')
                                                        if len(cells) >= 2:
                                                            # 두 번째 컬럼이 징계종류
                                                            dtype = cells[1].get_text(strip=True)
                                                            if not dtype or dtype == '-':
                                                                continue
                                                            # 포함 여부로 분류 (정직1월 → 정직, 감봉3월 → 감봉)
                                                            matched = False
                                                            # 출근정지 → 정직으로 분류
                                                            if "출근정지" in dtype:
                                                                discipline_counts["정직"] += 1
                                                                matched = True
                                                            else:
                                                                for key in ["파면", "해임", "정직", "감봉", "견책"]:
                                                                    if key in dtype:
                                                                        discipline_counts[key] += 1
                                                                        matched = True
                                                                        break
                                                            if not matched:
                                                                discipline_counts["기타"] += 1
                                                                org_other_types.append(dtype)
                                    except (requests.RequestException, ValueError, KeyError):
                                        pass

                                org["discipline_counts"] = discipline_counts
                                org["other_types"] = org_other_types
                                return org

                            completed = 0
                            eta_calc = ETACalculator(len(organ_list))
                            with ThreadPoolExecutor(max_workers=5) as executor:
                                futures = {executor.submit(fetch_discipline_detail, org): org for org in organ_list}

                                for future in as_completed(futures):
                                    if self._stop_event.is_set():
                                        return
                                    future.result()
                                    completed += 1

                                    if completed % 10 == 0 or completed == len(organ_list):
                                        progress = 30 + (completed / len(organ_list) * 70)
                                        self.progress_var.set(progress)
                                        eta_str = eta_calc.update(completed)
                                        self.status_var.set(f"징계 상세 조회 중... ({completed}/{len(organ_list)}) - {eta_str}")

                            # 총 징계 건수 기준 정렬 (많은 순)
                            organ_list.sort(key=lambda x: sum(x.get("discipline_counts", {}).values()), reverse=True)

                            total_disciplines = sum(sum(o.get("discipline_counts", {}).values()) for o in organ_list)
                            self.log(f"총 징계처분: {total_disciplines}건")

                            # 기타로 분류된 징계종류 집계
                            from collections import Counter
                            all_other_types = Counter()
                            for org in organ_list:
                                for ot in org.get("other_types", []):
                                    all_other_types[ot] += 1

                            if all_other_types:
                                self.log(f"[기타로 분류된 징계종류]")
                                for dtype, cnt in all_other_types.most_common(10):
                                    self.log(f"  - {dtype}: {cnt}건")

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())

            elif item_type == "integrity":
                # 청렴도 평가 결과: 기관별 연도별 등급 집계
                self.log("청렴도 평가 결과 API 호출 중...")
                api_url = f"{BASE_URL}/item/itemOrganListJung.json"

                request_body = {
                    "apbaId": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "quart": "",
                    "reportFormRootNo": root_no
                }

                headers = {
                    "Content-Type": "application/json;charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Origin": "https://www.alio.go.kr",
                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                }

                try:
                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        # 알리오 서버 응답 status 검증 (일부 항목은 서버에서 'fail' 반환)
                        if isinstance(data, dict) and data.get("status") and data.get("status") != "success":
                            srv_msg = data.get("message", "알 수 없음")
                            self.log(f"⚠ 알리오 서버 응답 실패: {srv_msg}")
                            self.log(f"  해당 공시항목(rootNo={root_no})은 알리오 사이트 자체 결함일 수 있습니다.")
                            self.log(f"  다른 공시항목을 시도하거나 잠시 후 다시 시도해 주세요.")
                            return

                        items = []
                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []
                                elif isinstance(inner_data, list):
                                    items = inner_data
                            elif 'organList' in data:
                                items = data.get('organList', []) or []

                        self.log(f"API 응답: {len(items)}건")

                        # 기관별로 그룹화 (최신 공시번호만 유지)
                        organ_grouped = {}

                        for item in items:
                            if self._stop_event.is_set():
                                return

                            apba_id = item.get("apbaId", "")
                            apba_na = item.get("apbaNa", "")
                            type_na = item.get("typeNa", "")
                            jidt_na = item.get("jidtNa", "")

                            # 필터링
                            if inst_type_filter != "전체" and inst_type_filter not in type_na:
                                continue
                            if dept_filter != "전체" and dept_filter not in jidt_na:
                                continue

                            # 지역 정보
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")

                            if region_filter != "전체" and region_filter not in area_na:
                                continue

                            if inst_name_filter and inst_name_filter.lower() not in apba_na.lower():
                                continue

                            disclosure_no = item.get("disclosureNo", "") or ""
                            submission_no = item.get("submissionNo", "") or ""

                            # 기관별로 최신 공시만 유지
                            if apba_id not in organ_grouped:
                                organ_grouped[apba_id] = {
                                    "apba_id": apba_id,
                                    "inst_name": apba_na,
                                    "inst_type": type_na,
                                    "dept": jidt_na,
                                    "region": area_na,
                                    "integrity_grades": {},
                                    "disclosure_no": disclosure_no,
                                    "submission_no": submission_no,
                                    "files_parsed": [],
                                }
                            else:
                                # 더 최신 공시번호로 업데이트 (submissionNo 기준)
                                if submission_no > organ_grouped[apba_id].get("submission_no", ""):
                                    organ_grouped[apba_id]["disclosure_no"] = disclosure_no
                                    organ_grouped[apba_id]["submission_no"] = submission_no

                        # 그룹화된 데이터를 리스트로 변환
                        for apba_id, org in organ_grouped.items():
                            organ_list.append(org)

                        self.log(f"기관별 그룹화: {len(organ_list)}개 기관")

                        # doc.html 파싱하여 청렴도 등급 추출 (병렬 처리)
                        if organ_list:
                            self.log("청렴도 등급 데이터 수집 중...")
                            all_years_found = set()

                            def fetch_integrity_grades(org):
                                """기관별 청렴도 등급 조회 - doc.html 파싱 (submissionNo 활용)"""
                                grades = {}
                                years_in_doc = []
                                # disclosureNo가 없으면 submissionNo 사용
                                doc_no = org.get("disclosure_no", "") or org.get("submission_no", "")

                                if not doc_no:
                                    return grades, years_in_doc

                                try:
                                    # doc.html URL 구성
                                    date_part = doc_no[:8] if len(doc_no) >= 8 else ""
                                    if date_part:
                                        year = date_part[:4]
                                        month = date_part[4:6]
                                        day = date_part[6:8]
                                        doc_url = f"{BASE_URL}/upload/disclosure/{year}/{month}/{day}/{doc_no}/doc.html"

                                        doc_resp = self.session.get(doc_url, timeout=30)
                                        if doc_resp.status_code == 200:
                                            soup = BeautifulSoup(doc_resp.text, 'html.parser')

                                            # 청렴도 테이블 찾기 (연도가 헤더에 있는 테이블)
                                            for table in soup.find_all('table', border='1'):
                                                headers = table.find_all('th')
                                                if headers:
                                                    header_texts = [th.get_text(strip=True) for th in headers]
                                                    # 연도 헤더 확인 (예: 2020년, 2021년, ...)
                                                    year_headers = [h for h in header_texts if h.endswith('년') and len(h) == 5]
                                                    if year_headers:
                                                        years_in_doc = [h.replace('년', '') for h in year_headers]
                                                        # 데이터 행 찾기
                                                        rows = table.find_all('tr')
                                                        for row in rows:
                                                            cells = row.find_all('td')
                                                            if cells and '청렴도' in cells[0].get_text():
                                                                # 첫 번째 셀은 구분, 나머지는 연도별 등급
                                                                for i, year_h in enumerate(year_headers):
                                                                    year_num = year_h.replace('년', '')
                                                                    if i + 1 < len(cells):
                                                                        grade = cells[i + 1].get_text(strip=True)
                                                                        # "해당없음" 등은 "-"로 변환
                                                                        if not grade or grade == "해당없음" or grade == "해당 없음":
                                                                            grade = "-"
                                                                        grades[year_num] = grade
                                                                break
                                                        if grades:
                                                            break
                                except (requests.RequestException, ValueError, KeyError):
                                    pass

                                return grades, years_in_doc

                            # 병렬 처리
                            eta_calc = ETACalculator(len(organ_list))
                            with ThreadPoolExecutor(max_workers=5) as executor:
                                futures = {executor.submit(fetch_integrity_grades, org): org for org in organ_list}
                                completed = 0
                                for future in as_completed(futures):
                                    org = futures[future]
                                    try:
                                        grades, years_in_doc = future.result()
                                        org["integrity_grades"] = grades
                                        all_years_found.update(years_in_doc)
                                    except Exception:
                                        pass
                                    completed += 1
                                    if completed % 20 == 0:
                                        eta_str = eta_calc.update(completed)
                                        self.log(f"  진행: {completed}/{len(organ_list)} - {eta_str}")
                                        self.progress_var.set(completed / len(organ_list) * 80)

                            self.log(f"청렴도 등급 수집 완료")

                        # 연도 정보 저장 (컬럼 헤더용) - 실제 데이터에서 발견된 연도 사용
                        if all_years_found:
                            self.integrity_years = sorted(list(all_years_found), reverse=True)[:5]
                        else:
                            # 데이터가 없으면 기본값
                            self.integrity_years = ["2024", "2023", "2022", "2021", "2020"]

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())

            elif item_type == "safety":
                # 사망자수(최근 5년) - 산업재해+안전사고 합산
                self.log("사망자수(최근 5년) API 호출 중...")
                api_url = f"{BASE_URL}/item/itemOrganListJung.json"

                request_body = {
                    "apbaId": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "quart": "",
                    "reportFormRootNo": root_no
                }

                headers = {
                    "Content-Type": "application/json;charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Origin": "https://www.alio.go.kr",
                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                }

                try:
                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        # 알리오 서버 응답 status 검증 (일부 항목은 서버에서 'fail' 반환)
                        if isinstance(data, dict) and data.get("status") and data.get("status") != "success":
                            srv_msg = data.get("message", "알 수 없음")
                            self.log(f"⚠ 알리오 서버 응답 실패: {srv_msg}")
                            self.log(f"  해당 공시항목(rootNo={root_no})은 알리오 사이트 자체 결함일 수 있습니다.")
                            self.log(f"  다른 공시항목을 시도하거나 잠시 후 다시 시도해 주세요.")
                            return

                        items = []
                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []
                                elif isinstance(inner_data, list):
                                    items = inner_data
                            elif 'organList' in data:
                                items = data.get('organList', []) or []

                        self.log(f"API 응답: {len(items)}건")

                        # 기관별로 그룹화 (최신 공시번호만 유지)
                        organ_grouped = {}

                        for item in items:
                            if self._stop_event.is_set():
                                return

                            apba_id = item.get("apbaId", "")
                            apba_na = item.get("apbaNa", "")
                            type_na = item.get("typeNa", "")
                            jidt_na = item.get("jidtNa", "")

                            # 필터링
                            if inst_type_filter != "전체" and inst_type_filter not in type_na:
                                continue
                            if dept_filter != "전체" and dept_filter not in jidt_na:
                                continue

                            # 지역 정보
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")

                            if region_filter != "전체" and region_filter not in area_na:
                                continue

                            if inst_name_filter and inst_name_filter.lower() not in apba_na.lower():
                                continue

                            disclosure_no = item.get("disclosureNo", "") or ""
                            submission_no = item.get("submissionNo", "") or ""

                            # 기관별로 최신 공시만 유지
                            if apba_id not in organ_grouped:
                                organ_grouped[apba_id] = {
                                    "apba_id": apba_id,
                                    "inst_name": apba_na,
                                    "inst_type": type_na,
                                    "dept": jidt_na,
                                    "region": area_na,
                                    "death_by_year": {},  # 연도별 사망자수 (산업재해+안전사고 합산)
                                    "disclosure_no": disclosure_no,
                                    "submission_no": submission_no,
                                    "files_parsed": [],
                                }
                            else:
                                # 더 최신 공시번호로 업데이트 (submissionNo 기준)
                                if submission_no > organ_grouped[apba_id].get("submission_no", ""):
                                    organ_grouped[apba_id]["disclosure_no"] = disclosure_no
                                    organ_grouped[apba_id]["submission_no"] = submission_no

                        # 그룹화된 데이터를 리스트로 변환
                        for apba_id, org in organ_grouped.items():
                            organ_list.append(org)

                        self.log(f"기관별 그룹화: {len(organ_list)}개 기관")

                        # doc.html 파싱하여 연도별 사망자수 추출 (병렬 처리)
                        if organ_list:
                            self.log("연도별 사망자수 데이터 수집 중...")
                            all_years_found = set()

                            def fetch_safety_data(org):
                                """기관별 연도별 사망자수 조회 - doc.html 파싱 (산업재해+안전사고 합산)"""
                                death_by_year = {}
                                years_in_doc = []
                                attach_files = []  # 첨부파일 목록
                                doc_no = org.get("disclosure_no", "") or org.get("submission_no", "")

                                if not doc_no:
                                    return death_by_year, years_in_doc, attach_files

                                try:
                                    # doc.html URL 구성
                                    date_part = doc_no[:8] if len(doc_no) >= 8 else ""
                                    if date_part:
                                        year = date_part[:4]
                                        month = date_part[4:6]
                                        day = date_part[6:8]
                                        doc_url = f"{BASE_URL}/upload/disclosure/{year}/{month}/{day}/{doc_no}/doc.html"

                                        doc_resp = self.session.get(doc_url, timeout=30)
                                        if doc_resp.status_code == 200:
                                            soup = BeautifulSoup(doc_resp.text, 'html.parser')

                                            # 첨부파일 추출 (안전경영책임보고서)
                                            for a_tag in soup.find_all('a', href=True):
                                                href = a_tag.get('href', '')
                                                if 'report_attach_down' in href:
                                                    # javascript:report_attach_down('파일명') 에서 파일명 추출
                                                    match = re.search(r"report_attach_down\(['\"](.+?)['\"]\)", href)
                                                    if match:
                                                        file_name = match.group(1)
                                                        attach_files.append({
                                                            "name": file_name,
                                                            "disclosure_no": doc_no
                                                        })

                                            # border="1" 테이블에서 사망자수 추출
                                            for table in soup.find_all('table', border='1'):
                                                # 모든 행에서 연도 헤더 찾기
                                                year_list = []  # 순서대로 연도 저장
                                                
                                                # thead 또는 첫 번째 tr에서 연도 헤더 찾기
                                                all_rows = table.find_all('tr')
                                                for row in all_rows:
                                                    ths = row.find_all('th')
                                                    for th in ths:
                                                        th_text = th.get_text(strip=True)
                                                        if th_text.endswith('년') and len(th_text) == 5:
                                                            year_num = th_text.replace('년', '')
                                                            if year_num not in year_list:
                                                                year_list.append(year_num)
                                                    if year_list:
                                                        break
                                                
                                                if not year_list:
                                                    continue
                                                
                                                years_in_doc = year_list.copy()
                                                
                                                # 산업재해 소계 행과 안전사고 사망자수 행 찾기
                                                industrial_by_year = {y: 0 for y in year_list}
                                                safety_by_year = {y: 0 for y in year_list}
                                                
                                                for row in all_rows:
                                                    cells = row.find_all(['td', 'th'])
                                                    cell_texts = [c.get_text(strip=True) for c in cells]
                                                    
                                                    # "소계" 행 (산업재해 사고사망자수)
                                                    if '소계' in cell_texts:
                                                        소계_idx = cell_texts.index('소계')
                                                        numbers_after = cell_texts[소계_idx+1:]
                                                        for y_idx, year_num in enumerate(year_list):
                                                            if y_idx < len(numbers_after):
                                                                try:
                                                                    industrial_by_year[year_num] = int(numbers_after[y_idx])
                                                                except (ValueError, TypeError):
                                                                    pass

                                                    # "안전사고" + "사망자수" 행
                                                    if '안전사고' in cell_texts and '사망자수' in cell_texts:
                                                        사망자수_idx = cell_texts.index('사망자수')
                                                        numbers_after = cell_texts[사망자수_idx+1:]
                                                        for y_idx, year_num in enumerate(year_list):
                                                            if y_idx < len(numbers_after):
                                                                try:
                                                                    safety_by_year[year_num] = int(numbers_after[y_idx])
                                                                except (ValueError, TypeError):
                                                                    pass
                                                
                                                # 연도별 합산
                                                for year_num in year_list:
                                                    ind = industrial_by_year.get(year_num, 0)
                                                    saf = safety_by_year.get(year_num, 0)
                                                    death_by_year[year_num] = ind + saf
                                                
                                                if death_by_year:
                                                    break
                                except Exception as e:
                                    pass

                                return death_by_year, years_in_doc, attach_files

                            # 병렬 처리
                            eta_calc = ETACalculator(len(organ_list))
                            with ThreadPoolExecutor(max_workers=5) as executor:
                                futures = {executor.submit(fetch_safety_data, org): org for org in organ_list}
                                completed = 0
                                for future in as_completed(futures):
                                    org = futures[future]
                                    try:
                                        death_by_year, years_in_doc, attach_files = future.result()
                                        org["death_by_year"] = death_by_year
                                        org["safety_files"] = attach_files  # 첨부파일 목록 저장
                                        all_years_found.update(years_in_doc)
                                    except Exception:
                                        pass
                                    completed += 1
                                    if completed % 20 == 0:
                                        eta_str = eta_calc.update(completed)
                                        self.log(f"  진행: {completed}/{len(organ_list)} - {eta_str}")
                                        self.progress_var.set(completed / len(organ_list) * 80)

                            self.log(f"사망자수 데이터 수집 완료")

                            # 연도 정보 저장 (컬럼 헤더용)
                            if all_years_found:
                                self.safety_years = sorted(list(all_years_found), reverse=True)[:5]
                            else:
                                self.safety_years = ["2024", "2023", "2022", "2021", "2020"]

                            # 총계 표시
                            total_death = sum(sum(o.get("death_by_year", {}).values()) for o in organ_list)
                            self.log(f"총 사망자수(5년 합계): {total_death}명")

                            # 최근 연도 사망자수 기준 정렬 (많은 순)
                            latest_year = self.safety_years[0] if self.safety_years else "2024"
                            organ_list.sort(key=lambda x: x.get("death_by_year", {}).get(latest_year, 0), reverse=True)

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())

            elif item_type == "envlaw":
                # 환경법규 위반현황 + v5.4.1 게시판형 항목 공통 처리
                _disp_name = self.item_var.get() or "게시판형 공시"
                self.log(f"{_disp_name} API 호출 중...")
                api_url = f"{BASE_URL}/item/itemOrganListJung.json"

                request_body = {
                    "apbaId": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "apbaNa": "",
                    "reportFormRootNo": root_no
                }

                headers = {
                    "Content-Type": "application/json;charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Origin": "https://www.alio.go.kr",
                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                }

                try:
                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        items = []
                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []

                        self.log(f"API 응답: {len(items)}개 기관")

                        # 각 기관별 위반건수 조회
                        for item in items:
                            if self._stop_event.is_set():
                                return

                            apba_id = item.get("apbaId", "")
                            apba_na = item.get("apbaNa", "")
                            type_na = item.get("typeNa", "")
                            jidt_na = item.get("jidtNa", "")

                            # 필터링
                            if inst_type_filter != "전체" and inst_type_filter not in type_na:
                                continue
                            if dept_filter != "전체" and dept_filter not in jidt_na:
                                continue

                            # 지역 정보
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")

                            if region_filter != "전체" and region_filter not in area_na:
                                continue

                            if inst_name_filter and inst_name_filter.lower() not in apba_na.lower():
                                continue

                            result = {
                                "apba_id": apba_id,
                                "inst_name": apba_na,
                                "inst_type": type_na,
                                "apba_type": item.get("apbaType", ""),
                                "dept": jidt_na,
                                "region": area_na,
                                "violation_count": 0,
                                "violations": [],
                            }

                            organ_list.append(result)

                        self.log(f"필터 적용 후: {len(organ_list)}개 기관")

                        # 각 기관별 위반 목록 조회 (병렬 처리)
                        if organ_list:
                            self.log(f"{_disp_name} 자료 건수 조회 중...")

                            def fetch_violation_list(org):
                                """기관별 자료 목록 조회 (envlaw + 게시판형 신규 항목 공용)"""
                                violations = []
                                apba_id = org.get("apba_id", "")
                                apba_type = org.get("apba_type", "")

                                try:
                                    # 위반 목록 API 호출 (itemReportListSusi.json)
                                    search_url = f"{BASE_URL}/item/itemReportListSusi.json"
                                    search_body = {
                                        "pageNo": 1,
                                        "apbaId": apba_id,
                                        "apbaType": apba_type,
                                        "reportFormRootNo": root_no,
                                        "search_word": "",
                                        "search_flag": "title",
                                        "bid_type": "",
                                        "enfc_istt": ""
                                    }
                                    search_resp = self.session.post(search_url, json=search_body, headers=headers, timeout=30)

                                    if search_resp.status_code == 200:
                                        search_data = search_resp.json()
                                        if 'data' in search_data and search_data['data']:
                                            result_list = search_data['data'].get('result', [])
                                            for v in result_list:
                                                violations.append({
                                                    "title": v.get("title", ""),
                                                    "idate": v.get("idate", ""),
                                                    "idx": v.get("idx", ""),
                                                    "submission_no": v.get("submissionNo", ""),
                                                    # v5.4.1: 게시판형 첨부파일 다운로드용 메타
                                                    "report_form_no": v.get("reportFormNo", ""),
                                                    "table_name": v.get("tableName", ""),
                                                    "idx_name": v.get("idxName", ""),
                                                    "bid_type": v.get("bidType", ""),
                                                    "disclosure_no": v.get("disclosureNo", ""),
                                                })
                                except requests.RequestException:
                                    pass

                                return violations

                            # 병렬 처리
                            eta_calc = ETACalculator(len(organ_list))
                            with ThreadPoolExecutor(max_workers=5) as executor:
                                futures = {executor.submit(fetch_violation_list, org): org for org in organ_list}
                                completed = 0
                                for future in as_completed(futures):
                                    org = futures[future]
                                    try:
                                        violations = future.result()
                                        org["violations"] = violations
                                        org["violation_count"] = len(violations)
                                    except Exception:
                                        pass
                                    completed += 1
                                    if completed % 10 == 0:
                                        eta_str = eta_calc.update(completed)
                                        self.log(f"  진행: {completed}/{len(organ_list)} - {eta_str}")
                                        self.progress_var.set(completed / len(organ_list) * 80)

                            # 위반 건수 기준 정렬 (많은 순)
                            organ_list.sort(key=lambda x: x.get("violation_count", 0), reverse=True)

                            total_violations = sum(o.get("violation_count", 0) for o in organ_list)
                            has_violations = len([o for o in organ_list if o.get("violation_count", 0) > 0])
                            self.log(f"{_disp_name}: 자료 있는 기관 {has_violations}개 / 총 {total_violations}건")

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())

            elif item_type == "audit":
                # 자체감사결과: 기관별 감사결과 건수 및 첨부파일
                self.log("자체감사결과 API 호출 중...")
                api_url = f"{BASE_URL}/item/itemOrganListJung.json"

                request_body = {
                    "apbaId": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "apbaNa": "",
                    "reportFormRootNo": root_no
                }

                headers = {
                    "Content-Type": "application/json;charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Origin": "https://www.alio.go.kr",
                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                }

                try:
                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        items = []
                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []

                        self.log(f"API 응답: {len(items)}개 기관")

                        # 각 기관별 데이터 수집
                        for item in items:
                            if self._stop_event.is_set():
                                return

                            apba_id = item.get("apbaId", "")
                            apba_na = item.get("apbaNa", "")
                            type_na = item.get("typeNa", "")
                            jidt_na = item.get("jidtNa", "")

                            # 필터링
                            if inst_type_filter != "전체" and inst_type_filter not in type_na:
                                continue
                            if dept_filter != "전체" and dept_filter not in jidt_na:
                                continue

                            # 지역 정보
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")

                            if region_filter != "전체" and region_filter not in area_na:
                                continue

                            if inst_name_filter and inst_name_filter.lower() not in apba_na.lower():
                                continue

                            result = {
                                "apba_id": apba_id,
                                "inst_name": apba_na,
                                "inst_type": type_na,
                                "apba_type": item.get("apbaType", ""),
                                "dept": jidt_na,
                                "region": area_na,
                                "audit_count": 0,
                                "audits": [],
                            }

                            organ_list.append(result)

                        self.log(f"필터 적용 후: {len(organ_list)}개 기관")

                        # 각 기관별 감사결과 목록 조회 (병렬 처리)
                        if organ_list:
                            self.log("감사결과 건수 조회 중...")

                            def fetch_audit_list(org):
                                """기관별 자체감사결과 목록 조회 (전체 페이지)"""
                                audits = []
                                apba_id = org.get("apba_id", "")
                                apba_type = org.get("apba_type", "")

                                try:
                                    # 감사결과 목록 API 호출 (itemReportListSusi.json)
                                    search_url = f"{BASE_URL}/item/itemReportListSusi.json"
                                    
                                    # 1페이지 먼저 조회하여 총 건수 파악
                                    search_body = {
                                        "pageNo": 1,
                                        "apbaId": apba_id,
                                        "apbaType": apba_type,
                                        "reportFormRootNo": root_no,
                                        "search_word": "",
                                        "search_flag": "title",
                                        "bid_type": "",
                                        "enfc_istt": ""
                                    }
                                    search_resp = self.session.post(search_url, json=search_body, headers=headers, timeout=30)

                                    if search_resp.status_code == 200:
                                        search_data = search_resp.json()
                                        if 'data' in search_data and search_data['data']:
                                            # 총 건수 및 페이지 정보 확인
                                            page_info = search_data['data'].get('page', {})
                                            total_count = page_info.get('totalCount', 0)
                                            total_page = page_info.get('totalPage', 1)
                                            
                                            # 1페이지 결과 추가
                                            result_list = search_data['data'].get('result', [])
                                            for v in result_list:
                                                audits.append({
                                                    "title": v.get("title", ""),
                                                    "idate": v.get("idate", ""),
                                                    "stDate": v.get("stDate", ""),
                                                    "idx": v.get("idx", ""),
                                                    "submission_no": v.get("submissionNo", ""),
                                                    "disclosure_no": v.get("disclosureNo", ""),
                                                })
                                            
                                            # 2페이지 이후 조회
                                            for page_no in range(2, total_page + 1):
                                                try:
                                                    search_body["pageNo"] = page_no
                                                    page_resp = self.session.post(search_url, json=search_body, headers=headers, timeout=30)
                                                    if page_resp.status_code == 200:
                                                        page_data = page_resp.json()
                                                        if 'data' in page_data and page_data['data']:
                                                            page_result = page_data['data'].get('result', [])
                                                            for v in page_result:
                                                                audits.append({
                                                                    "title": v.get("title", ""),
                                                                    "idate": v.get("idate", ""),
                                                                    "stDate": v.get("stDate", ""),
                                                                    "idx": v.get("idx", ""),
                                                                    "submission_no": v.get("submissionNo", ""),
                                                                    "disclosure_no": v.get("disclosureNo", ""),
                                                                })
                                                    time.sleep(0.5)
                                                except requests.RequestException:
                                                    pass
                                except requests.RequestException:
                                    pass

                                return audits

                            # 병렬 처리
                            eta_calc = ETACalculator(len(organ_list))
                            with ThreadPoolExecutor(max_workers=5) as executor:
                                futures = {executor.submit(fetch_audit_list, org): org for org in organ_list}
                                completed = 0
                                for future in as_completed(futures):
                                    org = futures[future]
                                    try:
                                        audits = future.result()
                                        org["audits"] = audits
                                        org["audit_count"] = len(audits)
                                    except Exception:
                                        pass
                                    completed += 1
                                    if completed % 10 == 0:
                                        eta_str = eta_calc.update(completed)
                                        self.log(f"  진행: {completed}/{len(organ_list)} - {eta_str}")
                                        self.progress_var.set(completed / len(organ_list) * 80)

                            # 감사결과 건수 기준 정렬 (많은 순)
                            organ_list.sort(key=lambda x: x.get("audit_count", 0), reverse=True)

                            total_audits = sum(o.get("audit_count", 0) for o in organ_list)
                            has_audits = len([o for o in organ_list if o.get("audit_count", 0) > 0])
                            self.log(f"감사결과 있는 기관: {has_audits}개 / 총 감사결과: {total_audits}건")

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())

            elif item_type == "mgmt_eval":
                # 경영실적 평가결과[공기업,준정부]: Susi API 2단계 호출
                self.log("경영실적 평가결과 API 호출 중...")
                api_url = f"{BASE_URL}/item/itemOrganListSusi.json"

                request_body = {
                    "apbaId": "",
                    "apbaNa": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "reportFormRootNo": root_no
                }

                headers = {
                    "Content-Type": "application/json;charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Origin": "https://www.alio.go.kr",
                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                }

                try:
                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        items = []
                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []
                                elif isinstance(inner_data, list):
                                    items = inner_data
                            elif 'organList' in data:
                                items = data.get('organList', []) or []

                        self.log(f"API 응답: {len(items)}개 기관")

                        # 각 기관별 데이터 수집
                        for item in items:
                            if self._stop_event.is_set():
                                return

                            apba_id = item.get("apbaId", "")
                            apba_na = item.get("apbaNa", "")
                            type_na = item.get("typeNa", "")
                            jidt_na = item.get("jidtNa", "")

                            # 필터링
                            if inst_type_filter != "전체" and inst_type_filter not in type_na:
                                continue
                            if dept_filter != "전체" and dept_filter not in jidt_na:
                                continue

                            # 지역 정보 (Susi API는 지역 미반환, public_inst_data에서 조회)
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")

                            if region_filter != "전체" and region_filter not in area_na:
                                continue

                            if inst_name_filter and inst_name_filter.lower() not in apba_na.lower():
                                continue

                            result = {
                                "apba_id": apba_id,
                                "inst_name": apba_na,
                                "inst_type": type_na,
                                "apba_type": item.get("apbaType", ""),
                                "dept": jidt_na,
                                "region": area_na,
                                "report_count": 0,
                                "reports": [],
                                "latest_title": "-",
                                "latest_idate": "-",
                            }

                            organ_list.append(result)

                        self.log(f"필터 적용 후: {len(organ_list)}개 기관")

                        # 각 기관별 보고서 목록 조회 (병렬 처리)
                        if organ_list:
                            self.log("보고서 목록 조회 중...")

                            def fetch_mgmt_eval_reports(org):
                                """기관별 경영실적 평가결과 보고서 목록 조회 (전체 페이지)"""
                                reports = []
                                apba_id = org.get("apba_id", "")
                                apba_type = org.get("apba_type", "")

                                try:
                                    search_url = f"{BASE_URL}/item/itemReportListSusi.json"

                                    # 1페이지 먼저 조회
                                    search_body = {
                                        "pageNo": 1,
                                        "apbaId": apba_id,
                                        "apbaType": apba_type,
                                        "reportFormRootNo": root_no,
                                        "search_word": "",
                                        "search_flag": "title",
                                        "bid_type": "",
                                        "enfc_istt": ""
                                    }
                                    search_resp = self.session.post(search_url, json=search_body, headers=headers, timeout=30)

                                    if search_resp.status_code == 200:
                                        search_data = search_resp.json()
                                        if 'data' in search_data and search_data['data']:
                                            page_info = search_data['data'].get('page', {})
                                            total_page = page_info.get('totalPage', 1)

                                            # 1페이지 결과 추가
                                            result_list = search_data['data'].get('result', [])
                                            for v in result_list:
                                                reports.append({
                                                    "title": v.get("title", ""),
                                                    "idate": v.get("idate", ""),
                                                    "disclosureNo": v.get("disclosureNo", ""),
                                                    "submissionNo": v.get("submissionNo", ""),
                                                    "reportFormNo": v.get("reportFormNo", ""),
                                                    "tableName": v.get("tableName", ""),
                                                    "idx": v.get("idx", ""),
                                                })

                                            # 2페이지 이후 조회
                                            for page_no in range(2, total_page + 1):
                                                try:
                                                    search_body["pageNo"] = page_no
                                                    page_resp = self.session.post(search_url, json=search_body, headers=headers, timeout=30)
                                                    if page_resp.status_code == 200:
                                                        page_data = page_resp.json()
                                                        if 'data' in page_data and page_data['data']:
                                                            page_result = page_data['data'].get('result', [])
                                                            for v in page_result:
                                                                reports.append({
                                                                    "title": v.get("title", ""),
                                                                    "idate": v.get("idate", ""),
                                                                    "disclosureNo": v.get("disclosureNo", ""),
                                                                    "submissionNo": v.get("submissionNo", ""),
                                                                    "reportFormNo": v.get("reportFormNo", ""),
                                                                    "tableName": v.get("tableName", ""),
                                                                    "idx": v.get("idx", ""),
                                                                })
                                                    time.sleep(0.5)
                                                except requests.RequestException:
                                                    pass
                                except requests.RequestException:
                                    pass

                                return reports

                            # 병렬 처리
                            eta_calc = ETACalculator(len(organ_list))
                            with ThreadPoolExecutor(max_workers=5) as executor:
                                futures = {executor.submit(fetch_mgmt_eval_reports, org): org for org in organ_list}
                                completed = 0
                                for future in as_completed(futures):
                                    org = futures[future]
                                    try:
                                        reports = future.result()
                                        org["reports"] = reports
                                        org["report_count"] = len(reports)
                                        if reports:
                                            # 최신 보고서 정보 (idate 기준 정렬)
                                            sorted_reports = sorted(reports, key=lambda x: x.get("idate", ""), reverse=True)
                                            org["latest_title"] = sorted_reports[0].get("title", "-")
                                            org["latest_idate"] = sorted_reports[0].get("idate", "-")
                                    except Exception:
                                        pass
                                    completed += 1
                                    if completed % 10 == 0:
                                        eta_str = eta_calc.update(completed)
                                        self.log(f"  진행: {completed}/{len(organ_list)} - {eta_str}")
                                        self.progress_var.set(completed / len(organ_list) * 80)

                            # 보고서 건수 기준 정렬 (많은 순)
                            organ_list.sort(key=lambda x: x.get("report_count", 0), reverse=True)

                            total_reports = sum(o.get("report_count", 0) for o in organ_list)
                            has_reports = len([o for o in organ_list if o.get("report_count", 0) > 0])
                            self.log(f"보고서 있는 기관: {has_reports}개 / 총 보고서: {total_reports}건")

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())

            elif item_type == "general":
                # 일반현황: 설립근거 파싱
                self.log("일반현황 API 호출 중...")
                api_url = f"{BASE_URL}/item/itemOrganListSusi.json"

                request_body = {
                    "apbaId": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "quart": "",
                    "reportFormRootNo": root_no
                }

                headers = {
                    "Content-Type": "application/json;charset=UTF-8",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Origin": "https://www.alio.go.kr",
                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                }

                try:
                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        items = []
                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []
                                elif isinstance(inner_data, list):
                                    items = inner_data
                            elif 'organList' in data:
                                items = data.get('organList', []) or []

                        self.log(f"API 응답: {len(items)}개 기관")

                        for item in items:
                            if self._stop_event.is_set():
                                return

                            apba_id = item.get("apbaId", "")
                            apba_na = item.get("apbaNa", "")
                            type_na = item.get("typeNa", "")
                            jidt_na = item.get("jidtNa", "")

                            # 필터링
                            if inst_type_filter != "전체" and inst_type_filter not in type_na:
                                continue
                            if dept_filter != "전체" and dept_filter not in jidt_na:
                                continue

                            # 지역 정보
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")

                            if region_filter != "전체" and region_filter not in area_na:
                                continue

                            if inst_name_filter and inst_name_filter.lower() not in apba_na.lower():
                                continue

                            result = {
                                "apba_id": apba_id,
                                "inst_name": apba_na,
                                "inst_type": type_na,
                                "apba_type": item.get("apbaType", ""),
                                "dept": jidt_na,
                                "region": area_na,
                                "purpose": "",  # 설립근거
                                "disclosure_no": item.get("disclosureNo", ""),
                                "submission_no": item.get("submissionNo", ""),
                            }

                            organ_list.append(result)

                        self.log(f"필터 적용 후: {len(organ_list)}개 기관")

                        # 각 기관별 설립근거 파싱 (병렬 처리)
                        if organ_list:
                            self.log("설립근거 데이터 수집 중...")

                            def fetch_purpose(org):
                                """기관별 설립근거 조회 - doc.html 파싱"""
                                purpose = ""
                                doc_no = org.get("disclosure_no", "") or org.get("submission_no", "")

                                if not doc_no:
                                    # API로 disclosureNo 조회
                                    try:
                                        susi_url = f"{BASE_URL}/item/itemReportListSusi.json"
                                        susi_body = {
                                            "pageNo": 1,
                                            "apbaId": org.get("apba_id", ""),
                                            "apbaType": org.get("apba_type", ""),
                                            "reportFormRootNo": root_no,
                                            "search_word": "",
                                            "search_flag": "title",
                                            "bid_type": "",
                                            "enfc_istt": ""
                                        }
                                        susi_resp = self.session.post(susi_url, json=susi_body, headers=headers, timeout=15)
                                        if susi_resp.status_code == 200:
                                            susi_data = susi_resp.json()
                                            if susi_data.get('status') == 'success':
                                                result_list = susi_data.get('data', {}).get('result', [])
                                                if result_list:
                                                    doc_no = str(result_list[0].get('disclosureNo', '') or result_list[0].get('submissionNo', ''))
                                    except requests.RequestException:
                                        pass

                                if not doc_no:
                                    return purpose, org_chart_url

                                try:
                                    # doc.html URL 구성
                                    date_part = doc_no[:8] if len(doc_no) >= 8 else ""
                                    if date_part:
                                        year = date_part[:4]
                                        month = date_part[4:6]
                                        day = date_part[6:8]
                                        doc_url = f"{BASE_URL}/upload/disclosure/{year}/{month}/{day}/{doc_no}/doc.html"

                                        doc_resp = self.session.get(doc_url, timeout=30)
                                        if doc_resp.status_code == 200:
                                            soup = BeautifulSoup(doc_resp.text, 'html.parser')

                                            # "설립근거" 텍스트 찾기
                                            for td in soup.find_all('td'):
                                                text = td.get_text(strip=True)
                                                if '설립근거' in text:
                                                    # 다음 행에서 설립근거 내용 추출
                                                    parent_table = td.find_parent('table')
                                                    if parent_table:
                                                        rows = parent_table.find_all('tr')
                                                        for i, row in enumerate(rows):
                                                            if '설립근거' in row.get_text():
                                                                # 다음 행 찾기
                                                                if i + 1 < len(rows):
                                                                    next_row = rows[i + 1]
                                                                    cells = next_row.find_all('td')
                                                                    if len(cells) >= 2:
                                                                        purpose = cells[-1].get_text(strip=True)
                                                                        if purpose:
                                                                            break
                                                    if purpose:
                                                        break

                                except (requests.RequestException, ValueError, KeyError):
                                    pass

                                return purpose

                            # 병렬 처리
                            eta_calc = ETACalculator(len(organ_list))
                            with ThreadPoolExecutor(max_workers=5) as executor:
                                futures = {executor.submit(fetch_purpose, org): org for org in organ_list}
                                completed = 0
                                for future in as_completed(futures):
                                    org = futures[future]
                                    try:
                                        purpose = future.result()
                                        org["purpose"] = purpose if purpose else "-"
                                    except Exception:
                                        org["purpose"] = "-"
                                    completed += 1
                                    if completed % 20 == 0:
                                        eta_str = eta_calc.update(completed)
                                        self.log(f"  진행: {completed}/{len(organ_list)} - {eta_str}")
                                        self.progress_var.set(completed / len(organ_list) * 80)

                            self.log(f"설립근거 수집 완료")

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())

            else:
                # 기존 항목별 공시 API 호출
                if item_type == "susi":
                    self.log("itemOrganListSusi.json API 호출 중...")
                    api_url = f"{BASE_URL}/item/itemOrganListSusi.json"
                else:
                    self.log("itemOrganListJung.json API 호출 중...")
                    api_url = f"{BASE_URL}/item/itemOrganListJung.json"

                request_body = {
                    "apbaId": "",
                    "apbaType": [],
                    "area": [],
                    "jidtDptm": [],
                    "quart": "",
                    "reportFormRootNo": root_no
                }

                self.log(f"URL: {api_url}")

                try:
                    headers = {
                        "Content-Type": "application/json;charset=UTF-8",
                        "Accept": "application/json, text/javascript, */*; q=0.01",
                        "X-Requested-With": "XMLHttpRequest",
                        "Origin": "https://www.alio.go.kr",
                        "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={root_no}"
                    }

                    resp = retry_request(self.session, "POST", api_url, json=request_body, headers=headers, timeout=30)
                    self.log(f"HTTP 상태: {resp.status_code}")

                    if resp.status_code == 200:
                        data = resp.json()
                        self.api_response_data = data

                        items = []

                        if isinstance(data, dict):
                            if 'data' in data and data['data'] is not None:
                                inner_data = data['data']
                                if isinstance(inner_data, dict):
                                    items = inner_data.get('organList', []) or []
                                    self.log(f"organList에서 {len(items)}개 발견")
                                elif isinstance(inner_data, list):
                                    items = inner_data
                            elif 'organList' in data:
                                items = data.get('organList', []) or []
                        elif isinstance(data, list):
                            items = data

                        # 데이터 파싱 및 필터링
                        for item in items:
                            if self._stop_event.is_set():
                                return

                            # 기관유형 필터링
                            type_na = item.get("typeNa", "")
                            if inst_type_filter != "전체":
                                if inst_type_filter not in type_na:
                                    continue

                            # 주무부처 필터링
                            jidt_na = item.get("jidtNa", "")
                            if dept_filter != "전체":
                                if dept_filter not in jidt_na:
                                    continue

                            # 기관명
                            apba_na = item.get("apbaNa", "")

                            # 지역 정보: ALIO 기관목록 API에서 가져온 정보 우선 사용
                            area_na = ""
                            if apba_na and apba_na in self.public_inst_data:
                                area_na = self.public_inst_data[apba_na].get("region", "")
                            if not area_na:
                                area_na = item.get("areaNa", "") or item.get("area", "")

                            # 지역 필터링
                            if region_filter != "전체":
                                if region_filter not in area_na:
                                    continue

                            # 기관명 필터링
                            if inst_name_filter:
                                if inst_name_filter.lower() not in apba_na.lower():
                                    continue

                            # files 필드 파싱
                            files_str = item.get("files", "")
                            files_parsed = parse_files_field(files_str)

                            # 공시기간 구성
                            crit_yyyy = item.get("critYyyy", "")
                            quart_na = item.get("quartNa", "")
                            period = f"{crit_yyyy} {quart_na}".strip()

                            result = {
                                "apba_id": item.get("apbaId", ""),
                                "inst_name": apba_na,
                                "inst_type": type_na,
                                "apba_type": item.get("apbaType", ""),
                                "dept": jidt_na,
                                "region": area_na,
                                "disclosure_no": item.get("disclosureNo", ""),
                                "submission_no": item.get("submissionNo", ""),
                                "report_form_no": item.get("reportFormNo", ""),
                                "files_raw": files_str,
                                "files_parsed": files_parsed,
                                "period": period,
                                "crit_yyyy": crit_yyyy,
                                "crit_quar": item.get("critQuar", ""),
                                "raw_data": item,
                            }

                            result["detail_url"] = f"{BASE_URL}/item/itemReportTerm.do?apbaId={result['apba_id']}&reportFormRootNo={root_no}"

                            organ_list.append(result)

                        self.log(f"필터 적용 후: {len(organ_list)}건")

                        # 기관별로 그룹화하여 최신 분기만 유지
                        organ_grouped = {}
                        for org in organ_list:
                            apba_id = org.get("apba_id", "")
                            crit_yyyy = org.get("crit_yyyy", "")
                            crit_quar = org.get("crit_quar", "")
                            # 정렬 키: 연도 + 분기 (예: "20251" > "20244")
                            sort_key = f"{crit_yyyy}{crit_quar}"

                            if apba_id not in organ_grouped:
                                organ_grouped[apba_id] = org
                            else:
                                existing_key = f"{organ_grouped[apba_id].get('crit_yyyy', '')}{organ_grouped[apba_id].get('crit_quar', '')}"
                                if sort_key > existing_key:
                                    organ_grouped[apba_id] = org

                        organ_list = list(organ_grouped.values())
                        self.log(f"기관별 최신 분기 정리: {len(organ_list)}개 기관")

                    else:
                        self.log(f"API 오류: HTTP {resp.status_code}")

                except Exception as e:
                    self.log(f"API 호출 오류: {str(e)}")
                    import traceback
                    self.log(traceback.format_exc())
            
            if not organ_list:
                self.log("")
                self.log("⚠ 검색 결과가 없습니다.")
                self.log("필터 조건을 변경해 보세요.")
            
            self.all_results = organ_list
            self.filtered_results = organ_list.copy()
            
            # 화면 표시
            is_rule = item_type == "rule"
            is_discipline = item_type == "discipline"
            is_integrity = item_type == "integrity"
            is_safety = item_type == "safety"
            is_envlaw = item_type == "envlaw"
            is_general = item_type == "general"
            is_audit = item_type == "audit"
            is_mgmt_eval = item_type == "mgmt_eval"

            # 열 헤더 설정 (항목별 다르게)
            if is_general:
                # 일반현황: 설립근거 표시
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text="설립근거")
                self.result_tree.heading("col7", text="")
                self.result_tree.heading("col8", text="")
                self.result_tree.heading("col9", text="")
                self.result_tree.heading("col10", text="")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=130)
                self.result_tree.column("dept", width=100)
                self.result_tree.column("col6", width=350)
                self.result_tree.column("col7", width=0)
                self.result_tree.column("col8", width=0)
                self.result_tree.column("col9", width=0)
                self.result_tree.column("col10", width=0)
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)
            elif is_discipline:
                # 징계처분 현황: 징계종류별 열 (파면/해임/정직/감봉/견책/기타)
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text="파면")
                self.result_tree.heading("col7", text="해임")
                self.result_tree.heading("col8", text="정직")
                self.result_tree.heading("col9", text="감봉")
                self.result_tree.heading("col10", text="견책")
                self.result_tree.heading("col11", text="기타")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=110)
                self.result_tree.column("dept", width=90)
                self.result_tree.column("col6", width=40, anchor="center")
                self.result_tree.column("col7", width=40, anchor="center")
                self.result_tree.column("col8", width=40, anchor="center")
                self.result_tree.column("col9", width=40, anchor="center")
                self.result_tree.column("col10", width=40, anchor="center")
                self.result_tree.column("col11", width=40, anchor="center")
                self.result_tree.column("col12", width=0)
            elif is_integrity:
                # 청렴도 평가 결과: 최근 5년 연도별 등급
                years = getattr(self, 'integrity_years', [str(datetime.now().year - i) for i in range(5)])
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text=years[0] if len(years) > 0 else "")
                self.result_tree.heading("col7", text=years[1] if len(years) > 1 else "")
                self.result_tree.heading("col8", text=years[2] if len(years) > 2 else "")
                self.result_tree.heading("col9", text=years[3] if len(years) > 3 else "")
                self.result_tree.heading("col10", text=years[4] if len(years) > 4 else "")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=110)
                self.result_tree.column("dept", width=90)
                self.result_tree.column("col6", width=70, anchor="center")
                self.result_tree.column("col7", width=70, anchor="center")
                self.result_tree.column("col8", width=70, anchor="center")
                self.result_tree.column("col9", width=70, anchor="center")
                self.result_tree.column("col10", width=70, anchor="center")
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)
            elif is_safety:
                # 사망자수(최근 5년): 연도별 표시
                years = getattr(self, 'safety_years', [str(datetime.now().year - i) for i in range(5)])
                self.result_tree.heading("select", text="선택")
                self.result_tree.heading("no", text="번호")
                self.result_tree.heading("apba_id", text="기관코드")
                self.result_tree.heading("inst_name", text="기관명")
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text=f"{years[0]}년" if len(years) > 0 else "")
                self.result_tree.heading("col7", text=f"{years[1]}년" if len(years) > 1 else "")
                self.result_tree.heading("col8", text=f"{years[2]}년" if len(years) > 2 else "")
                self.result_tree.heading("col9", text=f"{years[3]}년" if len(years) > 3 else "")
                self.result_tree.heading("col10", text=f"{years[4]}년" if len(years) > 4 else "")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("select", width=40, anchor="center")
                self.result_tree.column("no", width=45, anchor="center")
                self.result_tree.column("apba_id", width=70, anchor="center")
                self.result_tree.column("inst_name", width=180)
                self.result_tree.column("inst_type", width=110)
                self.result_tree.column("dept", width=90)
                self.result_tree.column("col6", width=70, anchor="center")
                self.result_tree.column("col7", width=70, anchor="center")
                self.result_tree.column("col8", width=70, anchor="center")
                self.result_tree.column("col9", width=70, anchor="center")
                self.result_tree.column("col10", width=70, anchor="center")
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)
            elif is_envlaw:
                # 환경법규 위반현황 + v5.4.1 게시판형: 자료 건수 표시
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text="자료건수")
                self.result_tree.heading("col7", text="")
                self.result_tree.heading("col8", text="")
                self.result_tree.heading("col9", text="")
                self.result_tree.heading("col10", text="")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=130)
                self.result_tree.column("dept", width=100)
                self.result_tree.column("col6", width=70, anchor="center")
                self.result_tree.column("col7", width=0)
                self.result_tree.column("col8", width=0)
                self.result_tree.column("col9", width=0)
                self.result_tree.column("col10", width=0)
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)
            elif is_audit:
                # 자체감사결과: 감사결과 건수 표시
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text="감사결과")
                self.result_tree.heading("col7", text="")
                self.result_tree.heading("col8", text="")
                self.result_tree.heading("col9", text="")
                self.result_tree.heading("col10", text="")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=130)
                self.result_tree.column("dept", width=100)
                self.result_tree.column("col6", width=70, anchor="center")
                self.result_tree.column("col7", width=0)
                self.result_tree.column("col8", width=0)
                self.result_tree.column("col9", width=0)
                self.result_tree.column("col10", width=0)
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)
            elif is_mgmt_eval:
                # 경영실적 평가결과: 보고서수, 최신보고서, 최신공시일
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text="지역")
                self.result_tree.heading("col7", text="보고서수")
                self.result_tree.heading("col8", text="최신보고서")
                self.result_tree.heading("col9", text="최신공시일")
                self.result_tree.heading("col10", text="")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=130)
                self.result_tree.column("dept", width=100)
                self.result_tree.column("col6", width=80)
                self.result_tree.column("col7", width=65, anchor="center")
                self.result_tree.column("col8", width=220)
                self.result_tree.column("col9", width=80, anchor="center")
                self.result_tree.column("col10", width=0)
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)
            elif is_rule:
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text="지역")
                self.result_tree.heading("col7", text="규정수")
                self.result_tree.heading("col8", text="규정분류")
                self.result_tree.heading("col9", text="")
                self.result_tree.heading("col10", text="")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=130)
                self.result_tree.column("col6", width=90)
                self.result_tree.column("col7", width=60)
                self.result_tree.column("col8", width=180)
                self.result_tree.column("col9", width=0)
                self.result_tree.column("col10", width=0)
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)
            else:
                self.result_tree.heading("inst_type", text="기관유형")
                self.result_tree.heading("dept", text="주무부처")
                self.result_tree.heading("col6", text="지역")
                self.result_tree.heading("col7", text="공시기간")
                self.result_tree.heading("col8", text="첨부파일")
                self.result_tree.heading("col9", text="공시번호")
                self.result_tree.heading("col10", text="")
                self.result_tree.heading("col11", text="")
                self.result_tree.heading("col12", text="")
                self.result_tree.column("inst_type", width=130)
                self.result_tree.column("col6", width=90)
                self.result_tree.column("col7", width=90)
                self.result_tree.column("col8", width=60)
                self.result_tree.column("col9", width=130)
                self.result_tree.column("col10", width=0)
                self.result_tree.column("col11", width=0)
                self.result_tree.column("col12", width=0)

            for idx, case in enumerate(self.filtered_results, 1):
                inst_type = case.get("inst_type", "")
                dept = case.get("dept", "")

                if is_general:
                    # 일반현황: 설립근거 표시
                    purpose = case.get("purpose", "-")
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        purpose[:50] + "..." if len(purpose) > 50 else purpose,
                        "", "", "", "", "", ""
                    )
                elif is_discipline:
                    # 징계처분 현황: 징계종류별 표시
                    dc = case.get("discipline_counts", {})
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        dc.get("파면", 0), dc.get("해임", 0), dc.get("정직", 0),
                        dc.get("감봉", 0), dc.get("견책", 0), dc.get("기타", 0), ""
                    )
                elif is_integrity:
                    # 청렴도 평가 결과: 연도별 등급 표시
                    grades = case.get("integrity_grades", {})
                    years = getattr(self, 'integrity_years', [str(datetime.now().year - i) for i in range(5)])
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        grades.get(years[0], "-") if len(years) > 0 else "-",
                        grades.get(years[1], "-") if len(years) > 1 else "-",
                        grades.get(years[2], "-") if len(years) > 2 else "-",
                        grades.get(years[3], "-") if len(years) > 3 else "-",
                        grades.get(years[4], "-") if len(years) > 4 else "-",
                        "", ""
                    )
                elif is_safety:
                    # 사망자수(최근 5년): 연도별 표시
                    death_by_year = case.get("death_by_year", {})
                    years = getattr(self, 'safety_years', ["2024", "2023", "2022", "2021", "2020"])
                    
                    # 안전하게 연도별 사망자수 가져오기
                    def safe_get_death(year_key):
                        try:
                            val = death_by_year.get(year_key, 0)
                            if isinstance(val, (int, float)):
                                return int(val)
                            return 0
                        except (ValueError, TypeError, KeyError):
                            return 0
                    
                    y0 = safe_get_death(years[0]) if len(years) > 0 else 0
                    y1 = safe_get_death(years[1]) if len(years) > 1 else 0
                    y2 = safe_get_death(years[2]) if len(years) > 2 else 0
                    y3 = safe_get_death(years[3]) if len(years) > 3 else 0
                    y4 = safe_get_death(years[4]) if len(years) > 4 else 0
                    
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        f"{y0}명" if y0 > 0 else "0",
                        f"{y1}명" if y1 > 0 else "0",
                        f"{y2}명" if y2 > 0 else "0",
                        f"{y3}명" if y3 > 0 else "0",
                        f"{y4}명" if y4 > 0 else "0",
                        "", ""
                    )
                elif is_envlaw:
                    # 환경법규 위반현황: 위반건수 표시
                    violation_count = case.get("violation_count", 0)
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        f"{violation_count}건" if violation_count > 0 else "0",
                        "", "", "", "", "", ""
                    )
                elif is_audit:
                    # 자체감사결과: 감사결과 건수 표시
                    audit_count = case.get("audit_count", 0)
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        f"{audit_count}건" if audit_count > 0 else "0",
                        "", "", "", "", "", ""
                    )
                elif is_mgmt_eval:
                    report_count = case.get("report_count", 0)
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        case.get("region", "") or "-",
                        f"{report_count}건" if report_count > 0 else "0",
                        case.get("latest_title", "-")[:30],
                        case.get("latest_idate", "-"),
                        "", "", ""
                    )
                elif is_rule:
                    rule_total = case.get("rule_total", 0)
                    rule_by_divis = case.get("rule_by_divis", {})
                    if rule_by_divis:
                        divis_parts = [f"{k}({v})" for k, v in rule_by_divis.items()]
                        divis_str = ", ".join(divis_parts[:4])
                        if len(divis_parts) > 4:
                            divis_str += "..."
                    else:
                        divis_str = "-"
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        case.get("region", "") or "-",
                        f"{rule_total}건" if rule_total else "0건",
                        divis_str, "", "", "", ""
                    )
                else:
                    files_count = len(case.get("files_parsed", []))
                    values = (
                        "[ ]", idx, case.get("apba_id", "") or "-", case.get("inst_name", ""),
                        inst_type, dept,
                        case.get("region", "") or "-",
                        case.get("period", "") or "-",
                        f"{files_count}개" if files_count else "-",
                        case.get("disclosure_no", "")[:20] if case.get("disclosure_no") else "-",
                        "", "", ""
                    )

                item_id = self.result_tree.insert("", tk.END, values=values)
                case["tree_id"] = item_id
            
            self.progress_var.set(100)
            self.update_selection_count()
            
            self.log("")
            self.log("=" * 60)
            self.log(f"기관 목록 수집 완료: {len(organ_list)}개 기관")
            self.log("기관을 선택 후 '공시내용수집'을 클릭하세요.")
            self.log("=" * 60)
            
            self.finish_process(True, True)
            
        except Exception as e:
            self.log(f"\n오류 발생: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            self.finish_process(False)
    
    def start_disclosure_fetch(self):
        """선택된 기관의 공시 내용 수집 (v5.4: 다중 항목 지원)"""
        selected_cases = []
        for item in self.result_tree.get_children():
            if item in self.selected_items:
                values = self.result_tree.item(item, "values")
                for case in self.filtered_results:
                    # apba_id 또는 기관명으로 매칭
                    if case.get("apba_id") == values[2] or case.get("inst_name") == values[3]:
                        selected_cases.append(case)
                        break

        if not selected_cases:
            messagebox.showwarning("알림", "수집할 기관을 선택해 주세요.")
            return

        # v5.4: 다중 항목 vs 단일 항목 분기
        item_infos = self.get_selected_item_infos()
        if len(item_infos) > 1:
            confirm_msg = (
                f"선택한 {len(selected_cases)}개 기관 × {len(item_infos)}개 공시항목 "
                f"(총 {len(selected_cases) * len(item_infos)}회)을 수집하시겠습니까?"
            )
        else:
            confirm_msg = f"선택한 {len(selected_cases)}개 기관의 공시 내용을 수집하시겠습니까?"

        if not messagebox.askyesno("확인", confirm_msg):
            return

        self._stop_event.clear()
        self._running_active = True
        self.search_btn.config(state=tk.DISABLED)
        self.detail_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress_var.set(0)

        if len(item_infos) > 1:
            thread = threading.Thread(
                target=self.fetch_disclosures_multi,
                args=(selected_cases, item_infos),
            )
        else:
            thread = threading.Thread(
                target=self.fetch_disclosures,
                args=(selected_cases,),
            )
        thread.daemon = True
        thread.start()

    def fetch_disclosures_multi(self, selected_cases, item_infos):
        """
        v5.4: 다중 공시항목 일괄 수집.
        폴더 구조: 저장경로 / ALIO_타임스탬프 / 항목명 / 기관명 / 파일들
        각 항목마다 fetch_disclosures를 호출 (직렬, 항목 단위 - 메모리 절약).
        """
        try:
            save_path = self.save_path_var.get()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_folder = os.path.join(save_path, f"ALIO_{timestamp}")
            os.makedirs(base_folder, exist_ok=True)

            self.log("")
            self.log("=" * 60)
            self.log(f"v5.4 다중 항목 일괄 수집 시작")
            self.log(f"  대상: {len(selected_cases)}개 기관 × {len(item_infos)}개 항목")
            self.log(f"  저장: {base_folder}")
            self.log("=" * 60)

            total_items = len(item_infos)
            for i, item_info in enumerate(item_infos):
                if self._stop_event.is_set():
                    break

                item_name = item_info.get("name") or item_info.get("rootNo", "(미상)")
                safe_item = sanitize_filename(item_name, max_len=40)
                item_folder = os.path.join(base_folder, safe_item)
                os.makedirs(item_folder, exist_ok=True)

                self.log("")
                self.log(f"━━━ [{i+1}/{total_items}] {item_name} ━━━")

                # 항목별 진행률 가중치 시작점
                self._multi_item_offset = (i / total_items) * 100
                self._multi_item_weight = 100 / total_items

                try:
                    self.fetch_disclosures(
                        selected_cases,
                        download_folder=item_folder,
                        override_item_info=item_info,
                        override_item_name=item_name,
                    )
                except Exception as e:
                    self.log(f"  ⚠ 항목 처리 오류: {e}")
                    continue

            # 정리
            self._multi_item_offset = None
            self._multi_item_weight = None
            self.progress_var.set(100)
            self.log("")
            self.log("=" * 60)
            self.log(f"다중 항목 일괄 수집 완료")
            self.log(f"저장 위치: {base_folder}")
            self.log("=" * 60)
            self.root.after(0, lambda: self.finish_process(success=True, enable_features=True))
        except Exception as e:
            self.log(f"오류: {e}")
            self.root.after(0, lambda: self.finish_process(success=False))
    
    def _build_condition_tag(self):
        """현재 검색조건 중 기본값이 아닌 것을 모아 파일명용 태그 문자열 반환"""
        parts = []
        inst_type = self.inst_type_var.get()
        if inst_type and inst_type != "전체":
            parts.append(sanitize_filename(inst_type, max_len=15))
        dept = self.dept_var.get()
        if dept and dept != "전체":
            parts.append(sanitize_filename(dept, max_len=15))
        region = self.region_var.get()
        if region and region != "전체":
            parts.append(sanitize_filename(region, max_len=10))
        inst_name = self.inst_name_var.get().strip()
        if inst_name:
            parts.append(sanitize_filename(inst_name, max_len=20))
        item_name = self.item_var.get()
        if item_name == "내부규정":
            rule_divis = self.rule_divis_var.get()
            if rule_divis and rule_divis != "전체":
                parts.append(sanitize_filename(rule_divis, max_len=15))
        return "_".join(parts)

    def fetch_disclosures(self, selected_cases, download_folder=None,
                          override_item_info=None, override_item_name=None):
        """
        공시 상세 내용 수집 및 파일 다운로드.

        v5.4 변경:
        - download_folder가 주어지면 그것을 사용 (다중 항목 처리 시 호출자가 사전 생성).
          미지정 시 기존 v5.3 방식 (ALIO_항목명_조건_타임스탬프).
        - override_item_info/override_item_name으로 다중 항목 처리 시
          현재 GUI 선택과 무관하게 특정 항목을 처리할 수 있게 함.
        """
        try:
            item_info = override_item_info or self.get_selected_item_info()
            if not item_info:
                self.log("")
                self.log("⚠ 공시항목이 선택되지 않았습니다.")
                self.log("   '공시항목 선택...' 버튼으로 항목을 먼저 선택해 주세요.")
                self.root.after(
                    0,
                    lambda: self.finish_process(success=False, enable_features=True),
                )
                return
            item_name = (
                override_item_name
                or self.item_var.get()
                or item_info.get("name", "(미상)")
            )
            item_type = item_info.get("type", "jung")
            root_no = item_info.get("rootNo")
            if not root_no:
                self.log("⚠ 선택된 공시항목 정보가 불완전합니다 (rootNo 없음). 항목을 다시 선택해 주세요.")
                self.root.after(
                    0,
                    lambda: self.finish_process(success=False, enable_features=True),
                )
                return

            if download_folder is None:
                save_path = self.save_path_var.get()
                safe_item_name = sanitize_filename(item_name, max_len=20)
                condition_tag = self._build_condition_tag()
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                folder_parts = ["ALIO", safe_item_name]
                if condition_tag:
                    folder_parts.append(condition_tag)
                folder_parts.append(timestamp)
                download_folder = os.path.join(save_path, "_".join(folder_parts))
            os.makedirs(download_folder, exist_ok=True)

            self.log("")
            self.log("=" * 60)
            self.log(f"공시 상세 내용 수집 시작: {item_name}")
            self.log("=" * 60)
            self.log(f"대상: {len(selected_cases)}개 기관")
            self.log(f"저장 폴더: {download_folder}")
            self.log("")

            total = len(selected_cases)
            success_count = 0
            pdf_count = 0
            file_count = 0

            # 내부규정인 경우 별도 처리
            if item_type == "rule":
                rule_divis = RULE_DIVIS_CODES.get(self.rule_divis_var.get(), "")

                for idx, case in enumerate(selected_cases):
                    if self._stop_event.is_set():
                        return

                    inst_name = case["inst_name"]
                    safe_inst_name = sanitize_filename(inst_name, max_len=30)

                    inst_folder = os.path.join(download_folder, safe_inst_name)
                    os.makedirs(inst_folder, exist_ok=True)

                    downloaded_files = []
                    inst_file_count = 0

                    # 내부규정 API 호출 (기관명으로 검색)
                    try:
                        rule_url = f"{BASE_URL}/occasional/findRuleList.json"

                        # 먼저 총 규정 개수 파악
                        params = {
                            "type": "apbaNa",
                            "word": inst_name,
                            "pageNo": 1,
                            "divis": rule_divis
                        }
                        resp = self.session.get(rule_url, params=params, timeout=30)

                        if resp.status_code != 200:
                            self.log(f"[{idx+1}/{total}] {inst_name} - API 오류")
                            continue

                        data = resp.json()
                        total_rule_cnt = data.get("data", {}).get("totalCnt", 0)

                        if total_rule_cnt == 0:
                            self.log(f"[{idx+1}/{total}] {inst_name} - 내부규정 없음")
                            continue

                        self.log(f"[{idx+1}/{total}] {inst_name} ({total_rule_cnt}건)")
                        self.status_var.set(f"수집 중: {inst_name} (0/{total_rule_cnt}건)")

                        page_no = 1
                        rule_count = 0

                        while True:
                            if self._stop_event.is_set():
                                return

                            if page_no > 1:
                                params["pageNo"] = page_no
                                resp = self.session.get(rule_url, params=params, timeout=30)
                                if resp.status_code != 200:
                                    break
                                data = resp.json()

                            rule_list = data.get("data", {}).get("result", [])
                            if not rule_list:
                                break

                            for rule in rule_list:
                                if self._stop_event.is_set():
                                    return

                                rule_count += 1
                                rule_title = rule.get("title", "")
                                rule_divis_name = rule.get("insdRuleDivis", "")
                                seq = rule.get("seq", "")

                                # 진행상황 업데이트
                                progress_pct = ((idx / total) + (rule_count / total_rule_cnt / total)) * 100
                                self.progress_var.set(progress_pct)
                                self.status_var.set(f"수집 중: {inst_name} ({rule_count}/{total_rule_cnt}건, {int(progress_pct)}%)")

                                if not seq:
                                    continue

                                # 상세 API 호출하여 파일 목록 가져오기
                                try:
                                    detail_url = f"{BASE_URL}/occasional/findRuleDtl.json?seq={seq}"
                                    detail_resp = self.session.get(detail_url, timeout=15)

                                    if detail_resp.status_code == 200:
                                        detail_data = detail_resp.json()
                                        b_files = detail_data.get("data", {}).get("bFiles", "")

                                        if b_files:
                                            file_entries = b_files.split(",")
                                            latest_file = None
                                            latest_file_no = 0

                                            for entry in file_entries:
                                                if "|" in entry:
                                                    file_no, file_name = entry.split("|", 1)
                                                    file_no = file_no.strip()
                                                    file_name = file_name.strip()

                                                    if not file_name.lower().endswith('.zip'):
                                                        try:
                                                            if int(file_no) > latest_file_no:
                                                                latest_file_no = int(file_no)
                                                                latest_file = (file_no, file_name)
                                                        except (ValueError, TypeError):
                                                            pass

                                            if latest_file:
                                                file_no, file_name = latest_file
                                                download_url = f"{BASE_URL}/download/rulefiledown.json?fileNo={file_no}"

                                                try:
                                                    file_resp = self.session.get(download_url, timeout=60)
                                                    if file_resp.status_code == 200 and len(file_resp.content) > 100:
                                                        safe_file_name = sanitize_filename(file_name, max_len=80)
                                                        divis_folder = os.path.join(inst_folder, rule_divis_name or "기타")
                                                        os.makedirs(divis_folder, exist_ok=True)

                                                        file_path = os.path.join(divis_folder, safe_file_name)

                                                        if os.path.exists(file_path):
                                                            base, ext = os.path.splitext(safe_file_name)
                                                            file_path = os.path.join(divis_folder, f"{base}_{file_count}{ext}")

                                                        with open(file_path, 'wb') as f:
                                                            f.write(file_resp.content)
                                                        downloaded_files.append(safe_file_name)
                                                        file_count += 1
                                                        inst_file_count += 1
                                                        self.log(f"    [{rule_count}/{total_rule_cnt}] [{rule_divis_name}] {file_name[:35]} ({len(file_resp.content)//1024}KB)")
                                                except Exception as e:
                                                    self.log(f"    [{rule_count}/{total_rule_cnt}] {file_name[:30]} (다운로드 오류)")

                                except Exception as e:
                                    pass

                                time.sleep(0.5)

                            # 다음 페이지
                            if page_no * 10 >= total_rule_cnt:
                                break
                            page_no += 1
                            time.sleep(0.5)

                        self.log(f"    ✓ 완료: {inst_file_count}개 파일 다운로드")

                    except Exception as e:
                        self.log(f"    → 오류: {str(e)[:40]}")

                    case["detail_fetched"] = True
                    case["downloaded_files"] = downloaded_files
                    case["save_folder"] = inst_folder

                    success_count += 1

                    progress = (idx + 1) / total * 100
                    self.progress_var.set(progress)
                    time.sleep(0.3)

            elif item_type == "discipline":
                # 징계처분 현황: 모든 분기별 파일 다운로드
                for idx, case in enumerate(selected_cases):
                    if self._stop_event.is_set():
                        return

                    inst_name = case["inst_name"]
                    apba_id = case.get("apba_id", "")
                    safe_inst_name = sanitize_filename(inst_name, max_len=30)

                    inst_folder = os.path.join(download_folder, f"{safe_inst_name}_{apba_id}")
                    os.makedirs(inst_folder, exist_ok=True)

                    downloaded_files = []
                    files_parsed = case.get("files_parsed", [])
                    disclosure_nos = case.get("disclosure_nos", [])

                    self.log(f"[{idx+1}/{total}] {inst_name} (분기: {len(disclosure_nos)}개, 첨부: {len(files_parsed)}개)")

                    # 각 공시번호별로 PDF 다운로드
                    for disc_no in disclosure_nos:
                        if self._stop_event.is_set():
                            return

                        try:
                            pdf_url = f"{BASE_URL}/download/pdf.json?disclosureNo={disc_no}"
                            pdf_resp = self.session.get(pdf_url, timeout=60)

                            if pdf_resp.status_code == 200:
                                content_type = pdf_resp.headers.get('Content-Type', '')

                                if 'pdf' in content_type.lower() or (len(pdf_resp.content) > 100 and pdf_resp.content[:4] == b'%PDF'):
                                    # 분기 정보 추출 시도
                                    pdf_filename = f"징계처분_{disc_no[-10:]}.pdf"
                                    pdf_path = os.path.join(inst_folder, pdf_filename)

                                    with open(pdf_path, 'wb') as f:
                                        f.write(pdf_resp.content)
                                    downloaded_files.append(pdf_filename)
                                    pdf_count += 1
                        except requests.RequestException:
                            pass

                        time.sleep(0.1)

                    # 첨부파일 다운로드
                    if files_parsed:
                        attach_folder = os.path.join(inst_folder, "첨부파일")

                        for fp in files_parsed:
                            if self._stop_event.is_set():
                                return

                            file_id = fp.get("id", "")
                            file_name = fp.get("name", "")
                            disc_no = fp.get("disclosure_no", "")
                            period = fp.get("period", "")

                            if not file_name or not file_id:
                                continue

                            try:
                                download_url = f"{BASE_URL}/download/file.json?f={file_id}&d={disc_no}"
                                file_resp = self.session.get(download_url, timeout=30)

                                if file_resp.status_code == 200 and len(file_resp.content) > 100:
                                    # 파일명에 분기 정보 추가
                                    safe_period = sanitize_filename(period, max_len=10)
                                    safe_file_name = sanitize_filename(file_name)
                                    if safe_period:
                                        final_name = f"{safe_period}_{safe_file_name}"
                                    else:
                                        final_name = safe_file_name

                                    # 첨부파일 폴더는 실제 다운로드 성공 시에만 생성
                                    os.makedirs(attach_folder, exist_ok=True)
                                    file_path = os.path.join(attach_folder, final_name)

                                    with open(file_path, 'wb') as f:
                                        f.write(file_resp.content)
                                    downloaded_files.append(final_name)
                                    file_count += 1
                            except requests.RequestException:
                                pass

                            time.sleep(0.5)

                    case["detail_fetched"] = True
                    case["downloaded_files"] = downloaded_files
                    case["save_folder"] = inst_folder

                    self.log(f"    → 다운로드: PDF {len([f for f in downloaded_files if f.endswith('.pdf')])}개, 첨부 {len([f for f in downloaded_files if not f.endswith('.pdf')])}개")

                    success_count += 1

                    progress = (idx + 1) / total * 100
                    self.progress_var.set(progress)
                    time.sleep(0.2)

            elif item_type == "envlaw":
                # v5.4.1: 환경법규 위반 + 게시판형 항목 — 기관 단위 병렬 처리
                _disp_envlaw = self.item_var.get() or "게시판형 공시"

                def _process_envlaw_inst(case_arg):
                    """기관 1개 envlaw/게시판형 처리 — 병렬 워커.
                    UI 업데이트 X, 결과 dict만 반환.
                    """
                    if self._stop_event.is_set():
                        return None
                    inst_name_l = case_arg["inst_name"]
                    apba_id_l = case_arg.get("apba_id", "")
                    violations_l = case_arg.get("violations", []) or []
                    if not violations_l:
                        return {"inst_name": inst_name_l, "case": case_arg,
                                "files": 0, "status": "no_data"}

                    safe_inst = sanitize_filename(inst_name_l, max_len=30)
                    inst_folder_l = os.path.join(download_folder, f"{safe_inst}_{apba_id_l}")
                    os.makedirs(inst_folder_l, exist_ok=True)

                    files_local = 0
                    # 외부 링크 누적 (입찰공고 B1030 같은 항목용)
                    external_links_per_v = {}

                    # 자료별 처리 — 첨부파일 또는 외부 링크 분기
                    for vi, vv in enumerate(violations_l, 1):
                        if self._stop_event.is_set():
                            break
                        has_rfn = bool(vv.get("report_form_no"))

                        if has_rfn:
                            # v5.4.1 게시판형: itemBoard{rfn}.do HTML 파싱
                            try:
                                attachments = fetch_board_attachment_list(
                                    self.session, apba_id_l, vv
                                )
                                if attachments:
                                    safe_title = sanitize_filename(vv.get("title", ""), max_len=40)
                                    sub_folder = os.path.join(inst_folder_l, f"{vi:02d}_{safe_title}")
                                    os.makedirs(sub_folder, exist_ok=True)
                                    for att in attachments:
                                        ok, _, _msg = download_board_attachment(
                                            self.session, att, sub_folder
                                        )
                                        if ok:
                                            files_local += 1
                                else:
                                    # 첨부 0개 → 외부 링크 시도 (입찰공고의 g2b 등)
                                    ext = fetch_board_external_links(
                                        self.session, apba_id_l, vv
                                    )
                                    if ext:
                                        external_links_per_v[vi] = ext
                            except (requests.RequestException, OSError):
                                pass
                        else:
                            # v5.3 환경법규 호환: doc.html 시도 (timeout 단축)
                            try:
                                sn = vv.get("submission_no", "") or vv.get("idx", "")
                                if sn and len(sn) >= 8:
                                    y, mo, d = sn[:4], sn[4:6], sn[6:8]
                                    doc_url = f"{BASE_URL}/upload/disclosure/{y}/{mo}/{d}/{sn}/doc.html"
                                    rr = self.session.get(doc_url, timeout=10)
                                    if rr.status_code == 200:
                                        safe_t = sanitize_filename(vv.get("title", ""), max_len=50)
                                        html_p = os.path.join(inst_folder_l, f"{vi:02d}_{safe_t}.html")
                                        with open(html_p, "w", encoding="utf-8") as f:
                                            f.write(rr.text)
                                        files_local += 1
                            except (requests.RequestException, OSError):
                                pass

                    # 자료 목록 텍스트 — 자료별 처리 후 작성 (외부 링크 포함)
                    list_file_l = os.path.join(inst_folder_l, "자료_목록.txt")
                    try:
                        with open(list_file_l, "w", encoding="utf-8") as f:
                            f.write(f"기관명: {inst_name_l}\n")
                            f.write(f"기관코드: {apba_id_l}\n")
                            f.write(f"항목명: {_disp_envlaw}\n")
                            f.write(f"자료건수: {len(violations_l)}건\n")
                            f.write("=" * 60 + "\n\n")
                            for vi, vv in enumerate(violations_l, 1):
                                f.write(f"[{vi}] {vv.get('title', '')}\n")
                                f.write(f"    등록일: {vv.get('idate', '')}\n")
                                f.write(f"    제출번호: {vv.get('submission_no', '')}\n")
                                # 외부 링크 (입찰공고 등)
                                if vi in external_links_per_v:
                                    f.write(f"    [외부 링크]\n")
                                    for el in external_links_per_v[vi]:
                                        f.write(f"      - {el['url']}\n")
                                        if el.get('text'):
                                            f.write(f"        ({el['text']})\n")
                                f.write("\n")
                        files_local += 1
                    except OSError:
                        pass

                    return {"inst_name": inst_name_l, "case": case_arg,
                            "files": files_local, "status": "ok",
                            "save_folder": inst_folder_l}

                # 병렬 실행 (max_workers=5 — 알리오 서버 부하 균형)
                eta_calc_e = ETACalculator(total)
                completed_e = 0
                with ThreadPoolExecutor(max_workers=5) as ex:
                    futures_e = {ex.submit(_process_envlaw_inst, c): c for c in selected_cases}
                    for fut in as_completed(futures_e):
                        if self._stop_event.is_set():
                            break
                        try:
                            res = fut.result()
                        except Exception as e:
                            self.log(f"  ⚠ 처리 오류: {e}")
                            continue
                        if res is None:
                            continue
                        completed_e += 1
                        if res["status"] == "no_data":
                            self.log(f"[{completed_e}/{total}] {res['inst_name']} - 자료 없음")
                        else:
                            self.log(f"[{completed_e}/{total}] {res['inst_name']} ({res['files']}개 파일)")
                            file_count += res["files"]
                            res["case"]["detail_fetched"] = True
                            res["case"]["save_folder"] = res["save_folder"]
                            success_count += 1

                        self.progress_var.set(completed_e / total * 100)
                        if completed_e % 10 == 0 or completed_e == total:
                            eta_str = eta_calc_e.update(completed_e)
                            self.log(f"  진행: {completed_e}/{total} - {eta_str}")

            elif item_type == "audit":
                # 자체감사결과: 감사결과 목록 및 첨부파일 다운로드
                for idx, case in enumerate(selected_cases):
                    if self._stop_event.is_set():
                        return

                    inst_name = case["inst_name"]
                    apba_id = case.get("apba_id", "")
                    audits = case.get("audits", [])
                    safe_inst_name = sanitize_filename(inst_name, max_len=30)

                    if not audits:
                        self.log(f"[{idx+1}/{total}] {inst_name} - 감사결과 없음")
                        continue

                    inst_folder = os.path.join(download_folder, f"{safe_inst_name}_{apba_id}")
                    os.makedirs(inst_folder, exist_ok=True)

                    self.log(f"[{idx+1}/{total}] {inst_name} ({len(audits)}건)")

                    downloaded_files = []

                    # 감사결과 목록을 텍스트 파일로 저장
                    list_file = os.path.join(inst_folder, "자체감사_결과_목록.txt")
                    with open(list_file, 'w', encoding='utf-8') as f:
                        f.write(f"기관명: {inst_name}\n")
                        f.write(f"기관코드: {apba_id}\n")
                        f.write(f"감사결과 건수: {len(audits)}건\n")
                        f.write("=" * 60 + "\n\n")

                        for a_idx, a in enumerate(audits, 1):
                            f.write(f"[{a_idx}] {a.get('title', '')}\n")
                            f.write(f"    등록일: {a.get('idate', '')}\n")
                            f.write(f"    제출번호: {a.get('submission_no', '') or a.get('disclosure_no', '')}\n")
                            f.write("\n")

                    file_count += 1

                    # 개별 감사결과 PDF 다운로드
                    total_audits = len(audits)
                    for a_idx, a in enumerate(audits, 1):
                        if self._stop_event.is_set():
                            return

                        disclosure_no = a.get('disclosure_no', '')
                        submission_no = a.get('submission_no', '')
                        title = a.get('title', '')
                        safe_title = sanitize_filename(title, max_len=50)

                        # 진행상황 표시
                        self.status_var.set(f"PDF 다운로드 중: {inst_name} ({a_idx}/{total_audits})")
                        sub_progress = (idx / total) + (a_idx / total_audits / total)
                        self.progress_var.set(sub_progress * 100)

                        if disclosure_no and submission_no:
                            # itemReportFiles.json에서 실제 파일명 조회
                            try:
                                files_url = f"{BASE_URL}/item/itemReportFiles.json?disclosureNo={disclosure_no}"
                                files_resp = self.session.get(files_url, timeout=30)
                                
                                if files_resp.status_code == 200:
                                    files_json = files_resp.json()
                                    
                                    # data 필드에서 파일 목록 추출
                                    files_data = files_json.get('data', []) if isinstance(files_json, dict) else files_json
                                    
                                    if files_data and len(files_data) > 0:
                                        for file_info in files_data:
                                            file_name = file_info.get('orcpFileNa', '')
                                            if file_name and file_name.lower().endswith('.pdf'):
                                                # PDF 파일 다운로드
                                                from urllib.parse import quote
                                                encoded_name = quote(file_name, safe='')
                                                download_url = f"{BASE_URL}/download/dfile.json?fileName={encoded_name}&submissionNo={submission_no}"
                                                
                                                pdf_resp = self.session.get(download_url, timeout=60)
                                                
                                                if pdf_resp.status_code == 200 and len(pdf_resp.content) > 100:
                                                    content_type = pdf_resp.headers.get('Content-Type', '')
                                                    if 'text/html' not in content_type.lower():
                                                        # 파일 저장
                                                        safe_file_name = sanitize_filename(file_name)
                                                        pdf_filename = f"{a_idx:02d}_{safe_file_name}"
                                                        pdf_path = os.path.join(inst_folder, pdf_filename)
                                                        
                                                        with open(pdf_path, 'wb') as f:
                                                            f.write(pdf_resp.content)
                                                        
                                                        downloaded_files.append(pdf_filename)
                                                        pdf_count += 1
                                                        break  # 첫 번째 PDF만 다운로드
                                                
                            except Exception as e:
                                pass

                        time.sleep(0.5)

                    case["detail_fetched"] = True
                    case["downloaded_files"] = downloaded_files
                    case["save_folder"] = inst_folder

                    pdf_files = len([f for f in downloaded_files if f.endswith('.pdf')])
                    other_files = len(downloaded_files) - pdf_files
                    self.log(f"    → 다운로드 완료: PDF {pdf_files}개, 기타 {other_files}개")
                    success_count += 1

                    progress = (idx + 1) / total * 100
                    self.progress_var.set(progress)
                    time.sleep(0.1)

            elif item_type == "mgmt_eval":
                # 경영실적 평가결과: 기관별 보고서 첨부파일 다운로드
                for idx, case in enumerate(selected_cases):
                    if self._stop_event.is_set():
                        return

                    inst_name = case["inst_name"]
                    apba_id = case.get("apba_id", "")
                    reports = case.get("reports", [])
                    safe_inst_name = sanitize_filename(inst_name, max_len=30)

                    if not reports:
                        self.log(f"[{idx+1}/{total}] {inst_name} - 보고서 없음")
                        continue

                    inst_folder = os.path.join(download_folder, f"{safe_inst_name}_{apba_id}")
                    os.makedirs(inst_folder, exist_ok=True)

                    self.log(f"[{idx+1}/{total}] {inst_name} ({len(reports)}건)")

                    downloaded_files = []

                    # 보고서 목록을 텍스트 파일로 저장
                    list_file = os.path.join(inst_folder, "경영실적_보고서_목록.txt")
                    with open(list_file, 'w', encoding='utf-8') as f:
                        f.write(f"기관명: {inst_name}\n")
                        f.write(f"기관코드: {apba_id}\n")
                        f.write(f"보고서 건수: {len(reports)}건\n")
                        f.write("=" * 60 + "\n\n")

                        for r_idx, r in enumerate(reports, 1):
                            f.write(f"[{r_idx}] {r.get('title', '')}\n")
                            f.write(f"    등록일: {r.get('idate', '')}\n")
                            f.write(f"    공시번호: {r.get('disclosureNo', '')}\n")
                            f.write(f"    제출번호: {r.get('submissionNo', '')}\n")
                            f.write("\n")

                    file_count += 1

                    # 개별 보고서 첨부파일 다운로드 (상세페이지 HTML 파싱 방식)
                    total_reports = len(reports)
                    for r_idx, r in enumerate(reports, 1):
                        if self._stop_event.is_set():
                            return

                        disclosure_no = r.get('disclosureNo', '')
                        report_form_no = r.get('reportFormNo', '') or root_no
                        table_name = r.get('tableName', '')
                        r_idx_val = r.get('idx', '')
                        title = r.get('title', '')

                        # 진행상황 표시
                        self.status_var.set(f"파일 다운로드 중: {inst_name} ({r_idx}/{total_reports})")
                        sub_progress = (idx / total) + (r_idx / total_reports / total)
                        self.progress_var.set(sub_progress * 100)

                        if not disclosure_no:
                            self.log(f"    [{r_idx}/{total_reports}] {title[:30]} - disclosureNo 없음, 건너뜀")
                            continue

                        try:
                            # Step 1: 상세페이지 HTML 조회
                            detail_url = (
                                f"{BASE_URL}/item/itemBoard{report_form_no}.do"
                                f"?disclosureNo={disclosure_no}"
                                f"&apbaId={apba_id}"
                                f"&nowcode={report_form_no}"
                                f"&reportFormNo={report_form_no}"
                                f"&table_name={table_name}"
                                f"&idx_name=BOARD_NO"
                                f"&idx={r_idx_val}"
                                f"&reportGbn=N"
                                f"&bid_type=0"
                            )
                            detail_resp = self.session.get(detail_url, timeout=30)

                            if detail_resp.status_code != 200:
                                self.log(f"    [{r_idx}/{total_reports}] 상세페이지 HTTP {detail_resp.status_code}")
                                continue

                            # Step 2: HTML에서 첨부파일 링크 파싱
                            # 패턴: <a href="/download/download.json?fileNo=2987679" ...>파일명.pdf</a>
                            file_links = re.findall(
                                r'href="/download/download\.json\?fileNo=(\d+)"[^>]*>([^<]+)</a>',
                                detail_resp.text
                            )

                            if not file_links:
                                self.log(f"    [{r_idx}/{total_reports}] {title[:30]} - 첨부파일 없음")
                                continue

                            # Step 3: 각 첨부파일 다운로드
                            for file_no, file_name in file_links:
                                if self._stop_event.is_set():
                                    return

                                file_name = file_name.strip()
                                download_url = f"{BASE_URL}/download/download.json?fileNo={file_no}"
                                file_resp = self.session.get(download_url, timeout=60)

                                if file_resp.status_code == 200 and len(file_resp.content) > 100:
                                    content_type = file_resp.headers.get('Content-Type', '')
                                    if 'text/html' not in content_type.lower():
                                        safe_file_name = sanitize_filename(file_name)
                                        save_name = f"{r_idx:02d}_{safe_file_name}"
                                        save_path_file = os.path.join(inst_folder, save_name)

                                        with open(save_path_file, 'wb') as f:
                                            f.write(file_resp.content)

                                        downloaded_files.append(save_name)
                                        pdf_count += 1
                                        self.log(f"    [{r_idx}/{total_reports}] {file_name} ({len(file_resp.content)//1024}KB)")
                                    else:
                                        self.log(f"    [{r_idx}/{total_reports}] {file_name} (HTML 응답, 건너뜀)")
                                else:
                                    self.log(f"    [{r_idx}/{total_reports}] {file_name} (HTTP {file_resp.status_code})")

                        except Exception as e:
                            self.log(f"    [{r_idx}/{total_reports}] {title[:30]} - 오류: {str(e)[:40]}")

                        time.sleep(0.5)

                    case["detail_fetched"] = True
                    case["downloaded_files"] = downloaded_files
                    case["save_folder"] = inst_folder

                    dl_count = len(downloaded_files)
                    self.log(f"    → 다운로드 완료: {dl_count}개 파일")
                    success_count += 1

                    progress = (idx + 1) / total * 100
                    self.progress_var.set(progress)
                    time.sleep(0.1)

            elif item_type == "safety":
                # 사망자수(최근 5년): 안전경영책임보고서 첨부파일 다운로드
                for idx, case in enumerate(selected_cases):
                    if self._stop_event.is_set():
                        return

                    inst_name = case["inst_name"]
                    apba_id = case.get("apba_id", "")
                    safety_files = case.get("safety_files", [])
                    disclosure_no = case.get("disclosure_no", "") or case.get("submission_no", "")
                    safe_inst_name = sanitize_filename(inst_name, max_len=30)

                    self.log(f"[{idx+1}/{total}] {inst_name} ({apba_id})")

                    inst_folder = os.path.join(download_folder, f"{safe_inst_name}_{apba_id}")
                    os.makedirs(inst_folder, exist_ok=True)

                    downloaded_files = []

                    # PDF 다운로드 (공시 PDF)
                    if disclosure_no:
                        pdf_url = f"{BASE_URL}/download/pdf.json?disclosureNo={disclosure_no}"
                        self.log(f"    → PDF 다운로드 시도...")

                        try:
                            pdf_resp = self.session.get(pdf_url, timeout=60)

                            if pdf_resp.status_code == 200:
                                content_type = pdf_resp.headers.get('Content-Type', '')
                                if 'pdf' in content_type.lower() or (len(pdf_resp.content) > 100 and pdf_resp.content[:4] == b'%PDF'):
                                    pdf_filename = f"사망자수_공시.pdf"
                                    pdf_path = os.path.join(inst_folder, pdf_filename)

                                    with open(pdf_path, 'wb') as f:
                                        f.write(pdf_resp.content)

                                    file_size = len(pdf_resp.content)
                                    self.log(f"    → PDF 저장 완료 ({file_size // 1024} KB)")
                                    downloaded_files.append(pdf_filename)
                                    pdf_count += 1
                        except Exception as e:
                            self.log(f"    → PDF 다운로드 실패: {str(e)[:30]}")

                    # 첨부파일(안전경영책임보고서) 다운로드
                    if safety_files:
                        self.log(f"    → 첨부파일 {len(safety_files)}개 다운로드 중...")
                        
                        # submissionNo 가져오기
                        submission_no = case.get("submission_no", "") or case.get("disclosure_no", "")
                        
                        for sf in safety_files:
                            if self._stop_event.is_set():
                                return

                            file_name = sf.get("name", "")

                            if not file_name:
                                continue

                            try:
                                # 첨부파일 다운로드 API: /download/dfile.json
                                from urllib.parse import quote
                                encoded_name = quote(file_name, safe='')
                                
                                download_url = f"{BASE_URL}/download/dfile.json?fileName={encoded_name}&submissionNo={submission_no}"
                                
                                file_resp = self.session.get(download_url, timeout=60)
                                
                                if file_resp.status_code == 200 and len(file_resp.content) > 100:
                                    # HTML 응답이 아닌지 확인
                                    content_type = file_resp.headers.get('Content-Type', '')
                                    if 'text/html' not in content_type.lower():
                                        safe_file_name = sanitize_filename(file_name)
                                        file_path = os.path.join(inst_folder, safe_file_name)

                                        with open(file_path, 'wb') as f:
                                            f.write(file_resp.content)

                                        downloaded_files.append(safe_file_name)
                                        file_count += 1
                                                
                            except Exception as e:
                                pass

                            time.sleep(0.5)

                    case["detail_fetched"] = True
                    case["downloaded_files"] = downloaded_files
                    case["save_folder"] = inst_folder

                    self.log(f"    → 다운로드 완료: {len(downloaded_files)}개 파일")
                    success_count += 1

                    progress = (idx + 1) / total * 100
                    self.progress_var.set(progress)
                    time.sleep(0.1)

            else:
                # 기존 공시 항목 처리 (복리후생비, 일반현황 등)
                for idx, case in enumerate(selected_cases):
                    if self._stop_event.is_set():
                        return

                    apba_id = case["apba_id"]
                    inst_name = case["inst_name"]
                    apba_type = case.get("apba_type", "")
                    safe_inst_name = sanitize_filename(inst_name, max_len=30)
                    disclosure_no = case.get("disclosure_no", "")
                    files_parsed = case.get("files_parsed", [])

                    self.log(f"[{idx+1}/{total}] {inst_name} ({apba_id})")

                    inst_folder = os.path.join(download_folder, f"{safe_inst_name}_{apba_id}")
                    os.makedirs(inst_folder, exist_ok=True)

                    attach_folder = os.path.join(inst_folder, "첨부파일")

                    downloaded_files = []
                    pdf_downloaded = False

                    # disclosureNo가 없는 경우 API에서 추출 시도
                    if not disclosure_no:
                        self.log(f"    → disclosureNo 없음, API 호출...")
                        # v5.4.1: 콤마 분리된 rootNos는 itemReportListSusi가 거부하므로
                        # 각 분기 rn별로 순차 시도
                        rn_candidates = [x.strip() for x in (root_no or "").split(",") if x.strip()]
                        for try_rn in rn_candidates:
                            try:
                                susi_url = f"{BASE_URL}/item/itemReportListSusi.json"
                                susi_body = {
                                    "pageNo": 1,
                                    "apbaId": apba_id,
                                    "apbaType": apba_type,
                                    "reportFormRootNo": try_rn,
                                    "search_word": "",
                                    "search_flag": "title",
                                    "bid_type": "",
                                    "enfc_istt": ""
                                }
                                susi_headers = {
                                    "Content-Type": "application/json;charset=UTF-8",
                                    "Accept": "application/json, text/javascript, */*; q=0.01",
                                    "X-Requested-With": "XMLHttpRequest",
                                    "Origin": "https://www.alio.go.kr",
                                    "Referer": f"https://www.alio.go.kr/item/itemOrganList.do?reportFormRootNo={try_rn}"
                                }
                                susi_resp = self.session.post(susi_url, json=susi_body, headers=susi_headers, timeout=15)
                                if susi_resp.status_code != 200:
                                    continue
                                susi_data = susi_resp.json()
                                if susi_data.get('status') != 'success':
                                    continue
                                result_list = susi_data.get('data', {}).get('result', []) or []
                                if not result_list:
                                    continue
                                disclosure_no = str(result_list[0].get('disclosureNo', '') or '')
                                if disclosure_no:
                                    self.log(f"    → disclosureNo 추출 ({try_rn}): {disclosure_no[:20]}...")
                                    break
                            except Exception as e:
                                self.log(f"    → API 호출 오류({try_rn}): {str(e)[:40]}")
                                continue

                    # PDF 다운로드
                    if disclosure_no:
                        pdf_url = f"{BASE_URL}/download/pdf.json?disclosureNo={disclosure_no}"
                        self.log(f"    → PDF 다운로드 시도...")

                        try:
                            pdf_resp = self.session.get(pdf_url, timeout=60)

                            if pdf_resp.status_code == 200:
                                content_type = pdf_resp.headers.get('Content-Type', '')

                                if 'pdf' in content_type.lower() or (len(pdf_resp.content) > 100 and pdf_resp.content[:4] == b'%PDF'):
                                    pdf_path = os.path.join(inst_folder, f"{safe_item_name}.pdf")
                                    with open(pdf_path, 'wb') as f:
                                        f.write(pdf_resp.content)
                                    pdf_downloaded = True
                                    pdf_count += 1
                                    downloaded_files.append(f"{safe_item_name}.pdf")
                                    self.log(f"    → PDF 저장 완료 ({len(pdf_resp.content)//1024} KB)")
                                else:
                                    try:
                                        pdf_data = pdf_resp.json()
                                        actual_url = pdf_data.get('url', '') or pdf_data.get('pdfUrl', '')
                                        if actual_url:
                                            if not actual_url.startswith('http'):
                                                actual_url = f"{BASE_URL}{actual_url}"
                                            pdf_resp2 = self.session.get(actual_url, timeout=60)
                                            if pdf_resp2.status_code == 200:
                                                pdf_path = os.path.join(inst_folder, f"{safe_item_name}.pdf")
                                                with open(pdf_path, 'wb') as f:
                                                    f.write(pdf_resp2.content)
                                                pdf_downloaded = True
                                                pdf_count += 1
                                                downloaded_files.append(f"{safe_item_name}.pdf")
                                                self.log(f"    → PDF 저장 완료 (간접)")
                                    except (requests.RequestException, json.JSONDecodeError, KeyError):
                                        self.log(f"    → PDF 응답 형식 불명")
                            else:
                                self.log(f"    → PDF 다운로드 실패: HTTP {pdf_resp.status_code}")
                        except Exception as e:
                            self.log(f"    → PDF 다운로드 오류: {str(e)[:40]}")
                    else:
                        self.log(f"    → disclosureNo 없음, PDF 생략")

                    # 첨부파일 다운로드
                    if files_parsed:
                        self.log(f"    → 첨부파일 {len(files_parsed)}개 발견")

                        for file_info in files_parsed:
                            file_id = file_info.get("id", "")
                            file_name = file_info.get("name", "")

                            if not file_name:
                                continue

                            if file_id and disclosure_no:
                                download_url = f"{BASE_URL}/download/file.json?f={file_id}&d={disclosure_no}"

                                try:
                                    file_resp = self.session.get(download_url, timeout=30)
                                    if file_resp.status_code == 200 and len(file_resp.content) > 100:
                                        # 첨부파일 폴더는 실제 다운로드 성공 시에만 생성
                                        os.makedirs(attach_folder, exist_ok=True)
                                        safe_file_name = sanitize_filename(file_name)
                                        file_path = os.path.join(attach_folder, safe_file_name)

                                        with open(file_path, 'wb') as f:
                                            f.write(file_resp.content)
                                        downloaded_files.append(safe_file_name)
                                        file_count += 1
                                        self.log(f"      - {file_name} ({len(file_resp.content)//1024} KB)")
                                except Exception as e:
                                    self.log(f"      - {file_name} (오류)")

                    case["detail_fetched"] = True
                    case["downloaded_files"] = downloaded_files
                    case["pdf_downloaded"] = pdf_downloaded
                    case["save_folder"] = inst_folder

                    success_count += 1

                    progress = (idx + 1) / total * 100
                    self.progress_var.set(progress)
                    time.sleep(0.3)

            self.log("")
            self.log("=" * 60)
            self.log("공시 내용 수집 완료!")
            self.log(f"성공: {success_count}/{total}개 기관")
            if item_type == "rule":
                self.log(f"내부규정 파일: {file_count}개")
            else:
                self.log(f"PDF: {pdf_count}개, 첨부파일: {file_count}개")
            self.log(f"저장 위치: {download_folder}")
            self.log("=" * 60)

            if item_type == "rule":
                msg = f"수집 완료!\n\n성공: {success_count}/{total}개 기관\n내부규정 파일: {file_count}개\n\n폴더를 열어보시겠습니까?"
            else:
                msg = f"수집 완료!\n\n성공: {success_count}/{total}개\nPDF: {pdf_count}개\n첨부파일: {file_count}개\n\n폴더를 열어보시겠습니까?"

            if messagebox.askyesno("완료", msg):
                import subprocess
                if sys.platform == "darwin":
                    subprocess.Popen(["open", download_folder])
                else:
                    subprocess.Popen(f'explorer "{download_folder}"', shell=True)

            self.finish_process(True, True)

        except Exception as e:
            self.log(f"\n오류 발생: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            self.finish_process(False)
    
    def export_to_excel(self):
        if not self.filtered_results:
            messagebox.showwarning("알림", "내보낼 데이터가 없습니다.")
            return
        
        save_path = self.save_path_var.get()
        os.makedirs(save_path, exist_ok=True)
        
        item_name = self.item_var.get()
        item_info = self.get_selected_item_info()
        item_type = item_info.get("type", "jung") if item_info else "jung"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ALIO 항목별공시"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 공시항목별 헤더 및 데이터 처리
        if item_type == "integrity":
            # 청렴도 평가 결과: 연도별 등급 포함
            years = getattr(self, 'integrity_years', [str(datetime.now().year - i) for i in range(5)])
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역"] + [f"{y}년" for y in years]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                grades = case.get("integrity_grades", {})
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                
                # 연도별 등급 입력
                for y_idx, year in enumerate(years):
                    grade = grades.get(year, "-")
                    cell = ws.cell(row=idx+1, column=7+y_idx, value=grade)
                    cell.border = thin_border
                    cell.alignment = cell_align_center
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            for i in range(len(years)):
                ws.column_dimensions[chr(ord('G') + i)].width = 10
        
        elif item_type == "discipline":
            # 징계처분 현황: 징계종류별 컬럼
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역", "파면", "해임", "정직", "감봉", "견책", "기타", "합계"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                dc = case.get("discipline_counts", {})
                total = sum(dc.values())
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                
                for d_idx, d_type in enumerate(["파면", "해임", "정직", "감봉", "견책", "기타"]):
                    cell = ws.cell(row=idx+1, column=7+d_idx, value=dc.get(d_type, 0))
                    cell.border = thin_border
                    cell.alignment = cell_align_center
                
                cell = ws.cell(row=idx+1, column=13, value=total)
                cell.border = thin_border
                cell.alignment = cell_align_center
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M']:
                ws.column_dimensions[col].width = 8
        
        elif item_type == "safety":
            # 사망자수(최근 5년): 연도별 컬럼
            years = getattr(self, 'safety_years', [str(datetime.now().year - i) for i in range(5)])
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역"] + [f"{y}년" for y in years]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                death_by_year = case.get("death_by_year", {})
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                
                # 연도별 사망자수 입력
                for y_idx, year in enumerate(years):
                    death_count = death_by_year.get(year, 0)
                    cell = ws.cell(row=idx+1, column=7+y_idx, value=death_count)
                    cell.border = thin_border
                    cell.alignment = cell_align_center
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            for i in range(len(years)):
                ws.column_dimensions[chr(ord('G') + i)].width = 10
        
        elif item_type == "envlaw":
            # 환경법규 위반현황
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역", "위반건수"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                cell = ws.cell(row=idx+1, column=7, value=case.get("violation_count", 0))
                cell.border = thin_border
                cell.alignment = cell_align_center
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 10
        
        elif item_type == "audit":
            # 자체감사결과
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역", "감사결과"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                cell = ws.cell(row=idx+1, column=7, value=case.get("audit_count", 0))
                cell.border = thin_border
                cell.alignment = cell_align_center
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 10
        
        elif item_type == "mgmt_eval":
            # 경영실적 평가결과
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역", "보고서수", "최신보고서", "최신공시일"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            for idx, case in enumerate(self.filtered_results, 1):
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                cell = ws.cell(row=idx+1, column=7, value=case.get("report_count", 0))
                cell.border = thin_border
                cell.alignment = cell_align_center
                ws.cell(row=idx+1, column=8, value=case.get("latest_title", "")).border = thin_border
                ws.cell(row=idx+1, column=9, value=case.get("latest_idate", "")).border = thin_border

            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 40
            ws.column_dimensions['I'].width = 12

        elif item_type == "rule":
            # 내부규정
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역", "규정수", "정관", "인사·복무·징계", "보수", "직제", "기타"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                rule_by_divis = case.get("rule_by_divis", {})
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                cell = ws.cell(row=idx+1, column=7, value=case.get("rule_total", 0))
                cell.border = thin_border
                cell.alignment = cell_align_center
                
                for r_idx, r_type in enumerate(["정관", "인사·복무·징계", "보수", "직제", "기타"]):
                    cell = ws.cell(row=idx+1, column=8+r_idx, value=rule_by_divis.get(r_type, 0))
                    cell.border = thin_border
                    cell.alignment = cell_align_center
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 8
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 8
            ws.column_dimensions['K'].width = 8
            ws.column_dimensions['L'].width = 8
        
        elif item_type == "general":
            # 일반현황
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역", "설립근거"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                ws.cell(row=idx+1, column=7, value=case.get("purpose", "")).border = thin_border
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 50
        
        else:
            # 기본 (복리후생비 등)
            headers = ["번호", "기관코드", "기관명", "기관유형", "주무부처", "지역", "공시기간", "첨부파일수", "공시번호", "상세URL"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for idx, case in enumerate(self.filtered_results, 1):
                files_count = len(case.get("files_parsed", []))
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=case.get("apba_id", "")).border = thin_border
                ws.cell(row=idx+1, column=3, value=case.get("inst_name", "")).border = thin_border
                ws.cell(row=idx+1, column=4, value=case.get("inst_type", "")).border = thin_border
                ws.cell(row=idx+1, column=5, value=case.get("dept", "")).border = thin_border
                ws.cell(row=idx+1, column=6, value=case.get("region", "")).border = thin_border
                ws.cell(row=idx+1, column=7, value=case.get("period", "")).border = thin_border
                ws.cell(row=idx+1, column=8, value=files_count).border = thin_border
                ws.cell(row=idx+1, column=9, value=case.get("disclosure_no", "")).border = thin_border
                ws.cell(row=idx+1, column=10, value=case.get("detail_url", "")).border = thin_border
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 22
            ws.column_dimensions['J'].width = 60
        
        ws.freeze_panes = "A2"
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_item_name = sanitize_filename(item_name, max_len=20)
        condition_tag = self._build_condition_tag()
        name_parts = ["ALIO", safe_item_name]
        if condition_tag:
            name_parts.append(condition_tag)
        name_parts.append(timestamp)
        filename = "_".join(name_parts) + ".xlsx"
        filepath = os.path.join(save_path, filename)
        
        wb.save(filepath)
        
        self.log(f"엑셀 저장 완료: {filepath}")
        messagebox.showinfo("완료", f"엑셀 파일이 저장되었습니다.\n\n{filepath}")
    
    def _cleanup(self):
        """종료 시 정리 작업"""
        self._stop_event.set()
        self._running_active = False
        # 설정 저장
        save_settings("alio_crawler", {"last_folder": self.save_path_var.get()})

    def stop_process(self):
        self._stop_event.set()
        self._running_active = False
        self.log("사용자에 의해 중지됨")
        self.finish_process(False)

    def finish_process(self, success=True, enable_features=False):
        self._stop_event.set()
        self._running_active = False
        self.search_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        
        if enable_features and self.all_results:
            self.detail_btn.config(state=tk.NORMAL)
            self.export_btn.config(state=tk.NORMAL)
            self.filter_btn.config(state=tk.NORMAL)
            self.clear_filter_btn.config(state=tk.NORMAL)
            self.select_all_btn.config(state=tk.NORMAL)
            self.deselect_all_btn.config(state=tk.NORMAL)
        
        self.status_var.set("완료" if success else "실패/중지")


def main():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl 라이브러리가 필요합니다. 설치 중...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    
    root = tk.Tk()
    app = ALIOCrawlerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()